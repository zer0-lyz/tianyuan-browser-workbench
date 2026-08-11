# -*- coding: utf-8 -*-
"""
资产基础法评估明细表「汇总链接恢复」工具（可复用）
====================================================
功能：
  1) 恢复所有被隐藏的科目/汇总表（sheet 设为可见）；
  2) 检测每张明细表的「账面价值/评估价值」列与「合计行」；
  3) 按 sheet 名称编号规律重建四层汇总链接：
       明细表合计 ──► 二级汇总（如 4-8固定资产汇总）
                   ──► 分类汇总（如 4-非流动资产汇总）
                   ──► 2-分类汇总
                   ──► 1-汇总表（万元换算 /10000）
  4) 逐链接对比「原静态值」与「链接计算值」，输出一致性报告。
用法：
  python 恢复汇总链接.py <明细表.xlsx> [输出目录]
"""
import sys, re, os
import openpyxl
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CI

AMOUNT_TOL = 0.01     # 金额容差（元）
RATE_TOL   = 0.01     # 比率容差（% 点）

# 需要按“净值”而非“原值”列取数的叶子表（原值/净值双列结构）
LEAF_COL_OVERRIDE = {
    '4-7-1投资性房地产': ('M', 'P'),   # 账面净值 / 评估净值（成本模式房屋）
    '4-7-2投资性房地产': ('M', 'N'),   # 账面价值 / 评估价值（公允价值模式房屋）
    '4-7-3投资性地产':   ('N', 'O'),   # 账面净值 / 评估价值（成本模式土地：L原始入账 M账面原值 N账面净值 O评估价值）
    '4-7-4投资性地产':   ('M', 'N'),   # 账面价值 / 评估价值（公允价值模式土地）
    '4-9-2在建（设备）': ('K', 'O'),   # 账面合计子列 / 评估合计子列
    '4-10生产性生物资产':('H', 'K'),   # 账面净值 / 评估净值
    '4-11油气资产':      ('J', 'M'),   # 账面净值 / 评估净值
}
# 分类汇总中取“净值”而不是“账面原值”的二级汇总（固定资产汇总 D=原值 E=净值 F=评估原值 G=评估净值）
CAT_CHILD_COL_OVERRIDE = {
    '4-8固定资产汇总': ('E', 'G'),
}

def norm(s):
    return re.sub(r'\s+', '', str(s)) if s is not None else ''

def to_num(v):
    """单元格数值容错：None/'' → None；数字 → float；文本数字 → float；
    其他文本（如 SpreadJS 计算错误占位符 '{"_calcError":"#VALUE!","_code":15}'）→ None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(',', '').strip()
        if s == '':
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None

def ssum(cells):
    """容错求和：返回 (合计, 异常占位符数)。非空且无法解析为数值的单元格计为异常。"""
    total, bad = 0.0, 0
    for v in cells:
        n = to_num(v)
        if n is None:
            if v not in (None, ''):
                bad += 1
        else:
            total += n
    return total, bad

def cell_ref(col, row):
    return f"{col}{row}"

def is_formula(v):
    return isinstance(v, str) and v.startswith('=')

def sheet_is_hidden(s):
    return s.sheet_state != 'visible'

# ---------------------------------------------------------------- 结构分析
def analyze(wbf, wbv):
    """返回 dict：{sheet: {bk, ev, hj, data_rows, kind}}，kind: leaf/sum2/cat/other"""
    info = {}
    CATS = {'3-流动资产汇总', '4-非流动资产汇总', '5-流动负债汇总', '6-非流动负债汇总'}
    SUM2 = {'3-1货币资金汇总表','3-2交易性金融资产汇总','3-8其他应收款汇总','3-9存货汇总',
            '4-6其他非流动金融资产汇总','4-7投资性房地产汇总','4-8固定资产汇总','4-9在建工程汇总',
            '4-13无形资产汇总','5-10其他应付款汇总表'}
    for ws in wbf.worksheets:
        name = ws.title
        d = {'bk': None, 'ev': None, 'hj': None, 'hj_rows': [], 'data_rows': [],
             'kind': 'other', 'rows_A': {}, 'rows_BC': {}}
        if name in ('1-汇总表', '2-分类汇总'):
            d['kind'] = 'top' if name == '1-汇总表' else 'class'
        elif name in CATS:
            d['kind'] = 'cat'
        elif name in SUM2:
            d['kind'] = 'sum2'
        elif name in ('设定信息', '00000000'):
            d['kind'] = 'other'
        else:
            d['kind'] = 'leaf'

        maxr, maxc = ws.max_row, ws.max_column
        # 行内容（按缓存值判断结构；公式文本取自 wbf）
        for r in range(1, maxr + 1):
            a = wbv[name].cell(r, 1).value
            if a is not None:
                d['rows_A'][r] = norm(a)
            b = wbv[name].cell(r, 2).value
            c = wbv[name].cell(r, 3).value
            if b is not None or c is not None:
                d['rows_BC'][r] = (norm(b), norm(c))
        # 合计行定位：**按内容动态识别，不依赖固定行号**。
        # 科目表(叶子表)数据条数不确定（同一模板各表 10~300+ 行不等，后续增删行也常见），
        # 合计行位置必然漂移，只能靠“A 列含‘合计’的最后一行”定位；
        # 汇总表位置相对固定，但同样按内容检测（B/C 列含“合计”）以兼容微调。
        # 1) 主判据：A 列（去空格后）含“合计”的最后一行
        hj_rows = [r for r, a in d['rows_A'].items() if '合计' in a]
        # 2) 汇总表“合计”可能标在 B/C 列（如 3-9存货汇总 C23=合 计）
        if not hj_rows and d['kind'] in ('cat', 'sum2'):
            hj_rows = [r for r, (b, c) in d['rows_BC'].items() if '合计' in (b or '') or '合计' in (c or '')]
        d['hj_rows'] = sorted(hj_rows)
        d['hj'] = hj_rows[-1] if hj_rows else None

        # 表头：4~8 行中找 账面价值/评估价值 列
        for r in range(4, 9):
            for c in range(1, maxc + 1):
                v = norm(wbf[name].cell(r, c).value)
                if '账面价值' in v and d['bk'] is None:
                    d['bk'] = CL(c)
                if '评估价值' in v and d['ev'] is None:
                    d['ev'] = CL(c)

        # 合并三列结构修复：部分表（3-9-x 存货系列等）"账面价值/评估价值"为合并单元格
        # 跨 数量/单价/金额 三列（模板原生表与 pandas 导出表均存在），R5 定位到的是合并区
        # 左上角（数量列），金额列实际在 R6 子表头的"金额"处。取 R6 精确等于"金额"的列
        # 覆盖 bk/ev；单列金额结构表（R6 无"金额"子表头）不受影响。
        amt_cols = [c for c in range(1, maxc + 1)
                    if norm(wbf[name].cell(6, c).value) == '金额']
        if amt_cols:
            if d['bk'] is not None and amt_cols[0] > CI(d['bk']):
                d['bk'] = CL(amt_cols[0])
            if d['ev'] is not None and len(amt_cols) > 1 and amt_cols[1] > CI(d['ev']):
                d['ev'] = CL(amt_cols[1])

        # 数据行：叶子表 A=序号数字；汇总表 A=编号 N-M(-K)
        if d['kind'] == 'leaf':
            d['data_rows'] = [r for r, a in d['rows_A'].items()
                              if re.fullmatch(r'\d+', a) and (d['hj'] is None or r < d['hj_rows'][0])]
        elif d['kind'] in ('cat', 'sum2'):
            d['data_rows'] = [r for r, a in d['rows_A'].items()
                              if re.fullmatch(r'\d+-\d+(-\d+)?', a)]

        # 叶子表列覆盖
        if d['kind'] == 'leaf' and name in LEAF_COL_OVERRIDE:
            d['bk'], d['ev'] = LEAF_COL_OVERRIDE[name]
        info[name] = d
    return info

# ---------------------------------------------------------------- 子表查找
def find_child(info, prefix, want_sum2=False):
    """按编号前缀找子表：优先 'xx汇总'，其次叶子表。返回 sheet 名或 None"""
    pat = re.compile(r'^' + re.escape(prefix) + r'(?![0-9\-])')
    cands = [n for n in info if pat.match(n)]
    if not cands:
        return None
    if want_sum2:
        for n in cands:
            if info[n]['kind'] == 'sum2':
                return n
        return None
    # 分类汇总行：优先二级汇总，否则叶子
    for n in cands:
        if info[n]['kind'] == 'sum2':
            return n
    for n in cands:
        if info[n]['kind'] == 'leaf':
            return n
    return cands[0]

# ---------------------------------------------------------------- 主流程
def main(path, outdir=None):
    wbf = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    info = analyze(wbf, wbv)
    name2ws = {ws.title: ws for ws in wbf.worksheets}

    # 1) 记录原始可见性：链接可写入隐藏表，但交付文件保持“原隐藏表仍隐藏”的状态
    #    （模板规范：用不到的表页应隐藏；00000000 为宏病毒残留文本页，同样保持隐藏）
    orig_states = {ws.title: ws.sheet_state for ws in wbf.worksheets}
    hidden_orig = [t for t, s in orig_states.items() if s != 'visible']

    # 有数据行但未识别到合计行的科目表 → 告警（防静默漏链）
    for sname, d in info.items():
        if d['kind'] == 'leaf' and d['data_rows'] and d['hj'] is None:
            print(f"  [警告] {sname!r} 有 {len(d['data_rows'])} 行明细但未识别到合计行，请人工检查该表合计行文字")

    # 记录原始值（供对比）
    originals = {}

    def snap(sheet, row, col):
        c = wbv[sheet].cell(row, CI(col) if isinstance(col, str) else col)
        originals[(sheet, row, col)] = c.value

    def put(sheet, col, row, formula):
        cell = name2ws[sheet].cell(row, CI(col) if isinstance(col, str) else col)
        # 已存在公式的单元格不覆盖（如 4-8固定资产汇总 已有链接）
        if is_formula(cell.value):
            return False
        # 原静态值为空 → 保持留空，不写入公式。
        # 原模板某列留空即该科目无此列概念（如 4-8-7固定资产清理 无账面原值/评估原值，
        # 其账面/评估净值在原模板 D/E 列），强行链接会把叶子表"账面价值"误挂到"账面原值"列。
        if wbv[sheet].cell(row, CI(col) if isinstance(col, str) else col).value in (None, ''):
            return False
        snap(sheet, row, col)
        cell.value = formula
        return True

    changes = []   # (sheet, cellname, formula, orig, expected)

    def add_change(sheet, col, row, formula, expected):
        changes.append((sheet, f"{col}{row}", formula, originals.get((sheet, row, col), '<已覆盖>'), expected))

    # ============ 第2层：二级汇总行 → 叶子表合计 ============
    for sname, d in info.items():
        if d['kind'] != 'sum2':
            continue
        for r in d['data_rows']:
            prefix = norm(name2ws[sname].cell(r, 1).value)
            child = find_child(info, prefix)
            if not child or info[child]['kind'] != 'leaf':
                continue
            cd = info[child]
            bsrc = cell_ref(cd['bk'], cd['hj']); esrc = cell_ref(cd['ev'], cd['hj'])
            fb = f"='{child}'!{bsrc}"; fe = f"='{child}'!{esrc}"
            okb = put(sname, 'D', r, fb); oke = put(sname, 'E', r, fe)
            bval = wbv[child][bsrc].value
            eval_ = wbv[child][esrc].value
            if okb:
                add_change(sname, 'D', r, fb, bval)
            if oke:
                add_change(sname, 'E', r, fe, eval_)
            if put(sname, 'F', r, f"=E{r}-D{r}"):
                add_change(sname, 'F', r, f"=E{r}-D{r}",
                           None if (bval is None or eval_ is None) else eval_ - bval)
            if put(sname, 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)'):
                rate = None if (bval is None or eval_ is None) else ("" if bval == 0 else (eval_ - bval) / abs(bval) * 100)
                add_change(sname, 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)', rate)

    # ============ 第3层：分类汇总行 → 二级汇总合计 / 叶子合计 ============
    for sname, d in info.items():
        if d['kind'] != 'cat':
            continue
        for r in d['data_rows']:
            prefix = norm(name2ws[sname].cell(r, 1).value)
            child = find_child(info, prefix)
            if not child:
                continue
            cd = info[child]
            if cd['kind'] == 'sum2':
                bcol, ecol = CAT_CHILD_COL_OVERRIDE.get(child, ('D', 'E'))
                bsrc = cell_ref(bcol, cd['hj']); esrc = cell_ref(ecol, cd['hj'])
                bval = wbv[child][bsrc].value; eval_ = wbv[child][esrc].value
            else:
                bsrc = cell_ref(cd['bk'], cd['hj']); esrc = cell_ref(cd['ev'], cd['hj'])
                bval = wbv[child][bsrc].value; eval_ = wbv[child][esrc].value
            fb = f"='{child}'!{bsrc}"; fe = f"='{child}'!{esrc}"
            if put(sname, 'D', r, fb):
                add_change(sname, 'D', r, fb, bval)
            if put(sname, 'E', r, fe):
                add_change(sname, 'E', r, fe, eval_)
            if put(sname, 'F', r, f"=E{r}-D{r}"):
                add_change(sname, 'F', r, f"=E{r}-D{r}",
                           None if (bval is None or eval_ is None) else eval_ - bval)
            if put(sname, 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)'):
                rate = None if (bval is None or eval_ is None) else ("" if bval == 0 else (eval_ - bval) / abs(bval) * 100)
                add_change(sname, 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)', rate)

    # ============ 第4层：2-分类汇总 → 分类汇总 ============
    cls = info['2-分类汇总']
    cat_rows = {}   # category -> {norm_name: row}
    for sname in ('3-流动资产汇总', '4-非流动资产汇总', '5-流动负债汇总', '6-非流动负债汇总'):
        cd = info[sname]
        cat_rows[sname] = {}
        for r in cd['data_rows']:
            nm = norm(name2ws[sname].cell(r, 3).value)
            cat_rows[sname][nm] = r

    def norm_name(s):
        s = norm(s)
        s = s.replace('其中：', '').replace('其中:', '')
        s = re.sub(r'[一二三四五六七八九十]、', '', s)
        s = re.sub(r'[、\s]', '', s)
        s = s.replace('合计', '').replace('总计', '')
        return s

    cls_orig = wbv['2-分类汇总']
    for r in range(7, cls['hj'] or 67 + 1):
        if r > 66:
            break
        a = cls_orig.cell(r, 1).value
        b = norm(cls_orig.cell(r, 2).value)
        if r in (7, 21, 40, 41, 55, 65, 66):
            continue   # 合计行单独处理
        if not re.fullmatch(r'\d+', str(a or '')):
            continue
        # 判断所属分类
        if 8 <= r <= 20:
            cat = '3-流动资产汇总'
        elif 22 <= r <= 39:
            cat = '4-非流动资产汇总'
        elif 42 <= r <= 54:
            cat = '5-流动负债汇总'
        elif 56 <= r <= 64:
            cat = '6-非流动负债汇总'
        else:
            continue
        rr = cat_rows[cat].get(b)
        if rr is None:
            nb = norm_name(b)
            for k, v in cat_rows[cat].items():
                if norm_name(k) == nb:
                    rr = v
                    break
        if rr is None:
            continue
        bval = wbv[cat].cell(rr, 4).value
        eval_ = wbv[cat].cell(rr, 5).value
        if put('2-分类汇总', 'C', r, f"='{cat}'!D{rr}"):
            add_change('2-分类汇总', 'C', r, f"='{cat}'!D{rr}", bval)
        if put('2-分类汇总', 'D', r, f"='{cat}'!E{rr}"):
            add_change('2-分类汇总', 'D', r, f"='{cat}'!E{rr}", eval_)
        if put('2-分类汇总', 'E', r, f"=D{r}-C{r}"):
            add_change('2-分类汇总', 'E', r, f"=D{r}-C{r}",
                       None if (bval is None or eval_ is None) else eval_ - bval)
        if put('2-分类汇总', 'F', r, f'=IF(C{r}=0,"",E{r}/ABS(C{r})*100)'):
            rate = None if (bval is None or eval_ is None) else ("" if bval == 0 else (eval_ - bval) / abs(bval) * 100)
            add_change('2-分类汇总', 'F', r, f'=IF(C{r}=0,"",E{r}/ABS(C{r})*100)', rate)

    # 2-分类汇总 合计行
    hjmap = {
        7:  ('3-流动资产汇总', None),
        21: ('4-非流动资产汇总', None),
        40: ('_sum', (7, 21)),
        41: ('5-流动负债汇总', None),
        55: ('6-非流动负债汇总', None),
        65: ('_sum', (41, 55)),
        66: ('_diff', (40, 65)),
    }
    for r, (src, arg) in hjmap.items():
        if src == '_sum':
            fC = f"=C{arg[0]}+C{arg[1]}"; fD = f"=D{arg[0]}+D{arg[1]}"
            expC = (cls_orig.cell(arg[0], 3).value or 0) + (cls_orig.cell(arg[1], 3).value or 0)
            expD = (cls_orig.cell(arg[0], 4).value or 0) + (cls_orig.cell(arg[1], 4).value or 0)
        elif src == '_diff':
            fC = f"=C{arg[0]}-C{arg[1]}"; fD = f"=D{arg[0]}-D{arg[1]}"
            expC = (cls_orig.cell(arg[0], 3).value or 0) - (cls_orig.cell(arg[1], 3).value or 0)
            expD = (cls_orig.cell(arg[0], 4).value or 0) - (cls_orig.cell(arg[1], 4).value or 0)
        else:
            cd = info[src]
            fC = f"='{src}'!D{cd['hj']}"; fD = f"='{src}'!E{cd['hj']}"
            expC = wbv[src].cell(cd['hj'], 4).value
            expD = wbv[src].cell(cd['hj'], 5).value
        if put('2-分类汇总', 'C', r, fC):
            add_change('2-分类汇总', 'C', r, fC, expC)
        if put('2-分类汇总', 'D', r, fD):
            add_change('2-分类汇总', 'D', r, fD, expD)
        if put('2-分类汇总', 'E', r, f"=D{r}-C{r}"):
            add_change('2-分类汇总', 'E', r, f"=D{r}-C{r}",
                       None if (expC is None or expD is None) else expD - expC)
        if put('2-分类汇总', 'F', r, f'=IF(C{r}=0,"",E{r}/ABS(C{r})*100)'):
            rate = None if (expC is None or expD is None) else ("" if expC == 0 else (expD - expC) / abs(expC) * 100)
            add_change('2-分类汇总', 'F', r, f'=IF(C{r}=0,"",E{r}/ABS(C{r})*100)', rate)

    # ============ 第5层：1-汇总表 → 2-分类汇总（万元） ============
    top = info['1-汇总表']
    cls_rows = {}   # norm_name -> row in 2-分类汇总
    for r in range(7, 67):
        b = norm(cls_orig.cell(r, 2).value)
        if b:
            cls_rows[norm_name(b)] = r
    top_ws = name2ws['1-汇总表']
    for r in range(8, 33):   # R8~R32，含最后的净资产行(R32)
        cname = norm(top_ws.cell(r, 3).value)
        if not cname:
            continue
        rr = cls_rows.get(norm_name(cname))
        if rr is None:
            continue
        bval = cls_orig.cell(rr, 3).value
        eval_ = cls_orig.cell(rr, 4).value
        expB = None if bval is None else bval / 10000
        expE = None if eval_ is None else eval_ / 10000
        if put('1-汇总表', 'D', r, f"='2-分类汇总'!C{rr}/10000"):
            add_change('1-汇总表', 'D', r, f"='2-分类汇总'!C{rr}/10000", expB)
        if put('1-汇总表', 'E', r, f"='2-分类汇总'!D{rr}/10000"):
            add_change('1-汇总表', 'E', r, f"='2-分类汇总'!D{rr}/10000", expE)
        if put('1-汇总表', 'F', r, f"=E{r}-D{r}"):
            add_change('1-汇总表', 'F', r, f"=E{r}-D{r}",
                       None if (expB is None or expE is None) else expE - expB)
        if put('1-汇总表', 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)'):
            rate = None if (expB is None or expE is None) else ("" if expB == 0 else (expE - expB) / abs(expB) * 100)
            add_change('1-汇总表', 'G', r, f'=IF(D{r}=0,"",F{r}/ABS(D{r})*100)', rate)

    # ============ 对比计算 ============
    def cmp_val(orig, exp):
        """orig=目标格原静态值, exp=链接来源格缓存值(None=源公式无缓存值,以Excel实算为准)"""
        if exp is None:
            return '源无缓存值'
        if isinstance(exp, str) and exp == '':
            return '一致' if orig in (0, '', None, '0') else '不一致'
        if isinstance(orig, str):
            try:
                orig = float(orig.replace(',', ''))
            except ValueError:
                return '不一致'
        if orig is None:
            return '一致(原为空)'
        diff = orig - exp
        # 金额与比率统一按 0.01 容差
        if abs(diff) <= AMOUNT_TOL:
            return '一致'
        return '不一致'

    report = []
    for sheet, cell, formula, orig, exp in changes:
        st = cmp_val(orig, exp)
        report.append({
            '表': sheet, '单元格': cell, '公式': formula,
            '原值': orig, '计算值': exp,
            '差异': None if orig is None or exp is None or isinstance(exp, str) else round((orig if not isinstance(orig, str) else 0) - exp, 6),
            '结果': st,
        })

    # ============ 叶子表合计验证（明细行合计 vs 现有合计行） ============
    # 结构：首个“合计”行为毛合计(SUM明细)，其后“减：xx准备”行，末行为净合计(毛-减值)
    leaf_check = []
    for sname, d in info.items():
        if d['kind'] != 'leaf' or not d['data_rows'] or not d['hj_rows'] or not d['bk']:
            continue
        first_hj, final_hj = d['hj_rows'][0], d['hj_rows'][-1]
        sum_b, bad_b = ssum(wbv[sname].cell(r, CI(d['bk'])).value for r in d['data_rows'])
        sum_e, bad_e = ssum(wbv[sname].cell(r, CI(d['ev'])).value for r in d['data_rows'])
        gross_b = wbv[sname].cell(first_hj, CI(d['bk'])).value
        gross_e = wbv[sname].cell(first_hj, CI(d['ev'])).value
        # 中间“减：”行对净合计的影响
        jz_b, _ = ssum(wbv[sname].cell(r, CI(d['bk'])).value for r in range(first_hj + 1, final_hj))
        jz_e, _ = ssum(wbv[sname].cell(r, CI(d['ev'])).value for r in range(first_hj + 1, final_hj))
        net_b = wbv[sname].cell(final_hj, CI(d['bk'])).value
        net_e = wbv[sname].cell(final_hj, CI(d['ev'])).value
        if gross_b is None or net_b is None:
            res_b = '源无缓存值'
        elif bad_b:
            res_b = '源含异常占位符'
        else:
            ok_gross_b = abs(sum_b - gross_b) <= AMOUNT_TOL
            ok_net_b = abs((to_num(gross_b) - jz_b) - to_num(net_b)) <= AMOUNT_TOL
            res_b = '一致' if ok_gross_b and ok_net_b else '不一致'
        if gross_e is None or net_e is None:
            res_e = '源无缓存值'
        elif bad_e:
            res_e = '源含异常占位符'
        else:
            ok_gross_e = abs(sum_e - gross_e) <= AMOUNT_TOL
            ok_net_e = abs((to_num(gross_e) - jz_e) - to_num(net_e)) <= AMOUNT_TOL
            res_e = '一致' if ok_gross_e and ok_net_e else '不一致'
        leaf_check.append({
            '表': sname, '毛合计行': first_hj, '净合计行': final_hj,
            '账面列': d['bk'], '评估列': d['ev'],
            '账面明细合计': round(sum_b, 2), '账面毛合计': gross_b if gross_b is not None else 0,
            '账面净合计': net_b if net_b is not None else 0, '账面减值': round(jz_b, 2),
            '评估明细合计': round(sum_e, 2), '评估毛合计': gross_e if gross_e is not None else 0,
            '评估净合计': net_e if net_e is not None else 0, '评估减值': round(jz_e, 2),
            '账面结果': res_b,
            '评估结果': res_e,
        })

    # 二级汇总/分类汇总 合计行自洽验证
    sum_check = []
    for sname, d in info.items():
        if d['kind'] not in ('cat', 'sum2') or not d['data_rows'] or d['hj'] is None:
            continue
        sum_b, bad_b = ssum(wbv[sname].cell(r, 4).value for r in d['data_rows'])
        sum_e, bad_e = ssum(wbv[sname].cell(r, 5).value for r in d['data_rows'])
        cur_b = wbv[sname].cell(d['hj'], 4).value
        cur_e = wbv[sname].cell(d['hj'], 5).value
        res_b = '源无缓存值' if cur_b is None else ('源含异常占位符' if bad_b else ('一致' if abs(sum_b - to_num(cur_b)) <= AMOUNT_TOL else '不一致'))
        res_e = '源无缓存值' if cur_e is None else ('源含异常占位符' if bad_e else ('一致' if abs(sum_e - to_num(cur_e)) <= AMOUNT_TOL else '不一致'))
        sum_check.append({
            '表': sname, '合计行': d['hj'],
            '账面行合计': round(sum_b, 2), '账面现合计': cur_b if cur_b is not None else 0,
            '评估行合计': round(sum_e, 2), '评估现合计': cur_e if cur_e is not None else 0,
            '账面结果': res_b,
            '评估结果': res_e,
        })

    # ============ 输出 ============
    outdir = outdir or os.path.dirname(os.path.abspath(path))
    os.makedirs(outdir, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    out_xlsx = os.path.join(outdir, base + '_链接恢复.xlsx')
    # 恢复原始 sheet 可见性（原隐藏表保持隐藏，链接公式保留在表内）
    for t, s in orig_states.items():
        if t in name2ws:
            name2ws[t].sheet_state = s
    wbf.save(out_xlsx)

    # ---- 对比报告 ----
    try:
        save_report(report, leaf_check, sum_check, hidden_orig, path, outdir, base)
    except Exception as e:
        print('报告写入失败:', e)

    return out_xlsx, report, leaf_check, sum_check, hidden_orig, info

def save_report(report, leaf_check, sum_check, restored, path, outdir, base):
    """生成《链接恢复对比报告.xlsx》"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '说明'
    CATS = {'3-流动资产汇总', '4-非流动资产汇总', '5-流动负债汇总', '6-非流动负债汇总'}
    SUM2 = {'3-1货币资金汇总表','3-2交易性金融资产汇总','3-8其他应收款汇总','3-9存货汇总',
            '4-6其他非流动金融资产汇总','4-7投资性房地产汇总','4-8固定资产汇总','4-9在建工程汇总',
            '4-13无形资产汇总','5-10其他应付款汇总表'}
    def layer(sheet):
        if sheet == '1-汇总表': return '第5层 总汇总表(万元)'
        if sheet == '2-分类汇总': return '第4层 分类汇总'
        if sheet in CATS: return '第3层 分类汇总表'
        if sheet in SUM2: return '第2层 二级汇总'
        return '其他'
    n_ok = sum(1 for x in report if x['结果'] == '一致')
    n_bad = sum(1 for x in report if x['结果'] == '不一致')
    n_warn = sum(1 for x in report if x['结果'] in ('一致(原为空)', '源无缓存值'))
    n_leaf_ok = sum(1 for x in leaf_check if x['账面结果'] == '一致' and x['评估结果'] == '一致')
    n_leaf_bad = sum(1 for x in leaf_check if x['账面结果'] == '不一致' or x['评估结果'] == '不一致')
    n_leaf_warn = len(leaf_check) - n_leaf_ok - n_leaf_bad
    n_sum_ok = sum(1 for x in sum_check if x['账面结果'] == '一致' and x['评估结果'] == '一致')
    n_sum_bad = sum(1 for x in sum_check if x['账面结果'] == '不一致' or x['评估结果'] == '不一致')
    rows = [
        ['资产基础法评估明细表 汇总链接恢复 对比报告'],
        [''],
        ['源文件', os.path.basename(path)],
        ['生成时间', __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        [''],
        ['一、恢复范围'],
        ['原隐藏表（已重建链接、仍保持隐藏）(张)', len(restored)],
        ['恢复的链接/公式(处)', len(report)],
        ['  其中：第5层 总汇总表(1-汇总表,万元)', sum(1 for x in report if x['表'] == '1-汇总表')],
        ['  其中：第4层 分类汇总(2-分类汇总)', sum(1 for x in report if x['表'] == '2-分类汇总')],
        ['  其中：第3层 分类汇总表(3/4/5/6-xx汇总)', sum(1 for x in report if layer(x['表']).startswith('第3层'))],
        ['  其中：第2层 二级汇总(如4-8固定资产汇总)', sum(1 for x in report if layer(x['表']).startswith('第2层'))],
        [''],
        ['二、一致性对比（链接计算值 vs 原静态值）'],
        ['对比处数', len(report)],
        ['一致', n_ok],
        ['原为空/源无缓存值(以Excel实算为准)', n_warn],
        ['不一致', n_bad],
        ['一致率', f'{n_ok/len(report)*100:.2f}%' if report else '-'],
        ['金额容差', f'±{AMOUNT_TOL} 元；比率容差 ±{RATE_TOL} 个百分点'],
        [''],
        ['三、叶子表合计验证（明细行合计 vs 表内合计行）'],
        ['验证叶子表数', len(leaf_check)],
        ['账面/评估均自洽', n_leaf_ok],
        ['源无缓存值', n_leaf_warn],
        ['存在差异', n_leaf_bad],
        [''],
        ['四、二级汇总/分类汇总 合计行自洽验证'],
        ['验证表数', len(sum_check)],
        ['账面/评估均自洽', n_sum_ok],
        ['存在差异', n_sum_bad],
        [''],
        ['五、说明'],
        ['1. 链接链条：明细表合计 → 二级汇总 → 分类汇总表 → 2-分类汇总 → 1-汇总表(÷10000 万元)。'],
        ['2. “不一致”即原静态值无法由底层明细重新汇总得到，需人工核实该科目原始数据。'],
        ['3. 固定资产类科目链接取“净值”列；原值/净值双列结构按模板列定义处理。'],
        ['4. 比率列(增值率)恢复为公式 =IF(D=0,"",F/D*100)，与原值(保留2位小数)在±0.01内视为一致。'],
        ['5. 模板自带 SUBTOTAL 小计(如4-8固定资产汇总)与勾稽校验公式(1-汇总表D34/E34)保留不动。'],
        ['6. 链接链条上各层公式已重建（含原隐藏表内），但交付文件中原隐藏表仍保持隐藏。'],
        ['7. “00000000”为旧版Excel宏病毒残留文本页(1998 XF.Classic)，非评估数据，建议删除。'],
    ]
    for r_ in rows:
        ws.append(r_)
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 30

    # 链接对比明细
    ws2 = wb.create_sheet('链接对比明细')
    ws2.append(['层级', '表', '单元格', '公式', '原值', '计算值', '差异', '结果'])
    for x in report:
        diff = x['差异']
        ws2.append([layer(x['表']), x['表'], x['单元格'], x['公式'],
                    x['原值'], x['计算值'], diff, x['结果']])
    # 公式列以“=”开头会被当作公式，强制为文本
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                c.data_type = 's'
    for col, w in zip('ABCDEFGH', (22, 24, 8, 46, 16, 16, 14, 8)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    # 叶子表合计验证
    ws3 = wb.create_sheet('叶子表合计验证')
    ws3.append(['表', '毛合计行', '净合计行', '账面列', '评估列',
                '账面明细合计', '账面毛合计', '账面减值', '账面净合计', '账面结果',
                '评估明细合计', '评估毛合计', '评估减值', '评估净合计', '评估结果'])
    for x in leaf_check:
        ws3.append([x['表'], x['毛合计行'], x['净合计行'], x['账面列'], x['评估列'],
                    x['账面明细合计'], x['账面毛合计'], x['账面减值'], x['账面净合计'], x['账面结果'],
                    x['评估明细合计'], x['评估毛合计'], x['评估减值'], x['评估净合计'], x['评估结果']])
    for col, w in zip('ABCDEFGHIJKLMNO', (26, 10, 10, 8, 8, 16, 16, 12, 16, 8, 16, 16, 12, 16, 8)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = 'A2'

    # 汇总表自洽
    ws4 = wb.create_sheet('汇总表合计自洽验证')
    ws4.append(['表', '合计行', '账面行合计', '账面现合计', '账面结果', '评估行合计', '评估现合计', '评估结果'])
    for x in sum_check:
        ws4.append([x['表'], x['合计行'], x['账面行合计'], x['账面现合计'], x['账面结果'],
                    x['评估行合计'], x['评估现合计'], x['评估结果']])
    for col, w in zip('ABCDEFGH', (26, 10, 18, 18, 10, 18, 18, 10)):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = 'A2'

    rep_path = os.path.join(outdir, base + '_链接恢复对比报告.xlsx')
    wb.save(rep_path)
    print('对比报告:', rep_path)
    return rep_path

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else '浙江省二轻集团有限责任公司-明细表.xlsx'
    outdir = sys.argv[2] if len(sys.argv) > 2 else None
    out_xlsx, report, leaf_check, sum_check, hidden_orig, info = main(src, outdir)
    n_ok = sum(1 for x in report if x['结果'] == '一致')
    n_bad = sum(1 for x in report if x['结果'] == '不一致')
    n_warn = sum(1 for x in report if x['结果'] in ('一致(原为空)', '源无缓存值'))
    print(f"输出文件: {out_xlsx}")
    print(f"原隐藏表 {len(hidden_orig)} 张（链接已重建，仍保持隐藏）: {hidden_orig}")
    print(f"共写入链接/公式 {len(report)} 处，一致 {n_ok}，原为空/源无缓存值 {n_warn}，不一致 {n_bad}")
    bad = [x for x in report if x['结果'] == '不一致']
    for x in bad[:20]:
        print(f"  [不一致] {x['表']}!{x['单元格']} 原值={x['原值']} 计算值={x['计算值']} 公式={x['公式']}")
    if leaf_check:
        badl = [x for x in leaf_check if x['账面结果'] == '不一致' or x['评估结果'] == '不一致']
        warnl = [x for x in leaf_check if x['账面结果'] == '源无缓存值' or x['评估结果'] == '源无缓存值']
        print(f"叶子表合计验证 {len(leaf_check)} 张，不一致 {len(badl)}，源无缓存值 {len(warnl)}")
        for x in badl[:20]:
            print(f"  [合计不一致] {x['表']} 账面 {x['账面明细合计']} vs 毛{x['账面毛合计']} 净{x['账面净合计']} / 评估 {x['评估明细合计']} vs 毛{x['评估毛合计']} 净{x['评估净合计']}")
