# -*- coding: utf-8 -*-
"""
评估明细表公式恢复 Skill 自动校对工具（三步自动化校验）
====================================================
输入:
  A = 系统原始导出基准文件（汇总单元格为静态数值，无恢复公式）
  B = 经过 Skill 公式恢复之后的输出文件
约束:
  假设 A、B 两份文件明细行未被修改；若检测到明细行内容不一致 → 直接报告比对无效，终止数值校验。

执行三步自动化校验:
  Step 1 单元格完整性校验: 对比 A/B 全部 sheet，表头/明细/备注/空行等所有非汇总行
          单元格文本、数值不得改动，改动记入异常清单。
  Step 2 汇总行公式合法性校验: ①应处理汇总行必须生成 Excel 公式；
          ②非汇总行不能新增公式；③公式不得产生 #REF! / #NAME? / #VALUE! 错误。
  Step 3 数值一致性比对: key = sheet+行号+科目编码；A 静态值(基准) vs B 公式计算值，
          浮点容差 0.01，|差|<=0.01 判定一致；不一致输出异常清单。

业务策略（不误报）:
  - 空汇总行（A 原值为空）: Skill 策略留空不写公式 → 不要求生成公式；
  - 已有旧公式行（A 中已是公式，如 4-8固定资产汇总 SUBTOTAL、1-汇总表 D34/E34 勾稽校验）:
    B 保留原公式，不算新增异常；
  - 源文件 SpreadJS 计算错误占位符（{"_calcError":"#VALUE!",...}）: 标记"源异常占位符"，
    不参与数值一致性（无法解析）。

用法:
  python audit_restore.py <A基准文件.xlsx> <B输出文件.xlsx> [报告输出目录]
"""
import sys
import os
import re
import datetime
import openpyxl
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CI

SKILL_SCRIPTS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'scripts')
if SKILL_SCRIPTS not in sys.path:
    sys.path.insert(0, SKILL_SCRIPTS)
import restore_links as RL  # 复用 analyze 结构分析

AMOUNT_TOL = 0.01   # 数值容差
ERR_MARKERS = ('#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#N/A', '#NULL!', '#NUM!')
PLACEHOLDER_RE = re.compile(r'\{"_calcError":"([^"]*)"')

CELL_REF = re.compile(r"'([^']+)'!(\$?[A-Z]+)(\$?\d+)")
LOCAL_REF = re.compile(r'(?<![A-Z0-9$])(\$?[A-Z]{1,2})(\$?\d+)(?![A-Z0-9])')


def norm(s):
    return re.sub(r'\s+', '', str(s)) if s is not None else ''


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(',', '').strip()
        if s == '':
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def placeholder_code(v):
    """返回 SpreadJS 占位符的错误码，如 #VALUE!；非占位符返回 None"""
    if isinstance(v, str):
        m = PLACEHOLDER_RE.search(v)
        if m:
            return m.group(1)
    return None


class Auditor:
    def __init__(self, path_a, path_b):
        self.path_a = path_a
        self.path_b = path_b
        self.wa = openpyxl.load_workbook(path_a, data_only=False)   # A 公式版（含旧公式）
        self.wav = openpyxl.load_workbook(path_a, data_only=True)   # A 值版（静态值/缓存值）
        self.wb = openpyxl.load_workbook(path_b, data_only=False)   # B 公式版
        self.wbv = openpyxl.load_workbook(path_b, data_only=True)   # B 值版
        self._eval_cache = {}
        self.summary_rows = None   # {(sheet,row): 科目编码/名称}
        self.summary_cells = set()  # {(sheet,row,col)} 应处理汇总单元格

    # ---------------- 汇总行识别（复用 restore_links 结构分析） ----------------
    def identify(self):
        info = RL.analyze(self.wa, self.wav)
        self.info = info
        summary_rows = {}     # (sheet,row) -> (科目编码, 科目名称, is_total)
        summary_cells = set()

        def add_row(sheet, row, code, name, is_total=False):
            summary_rows[(sheet, row)] = (code, name, is_total)

        for sname, d in info.items():
            ws = self.wa[sname]
            if d['kind'] in ('sum2', 'cat'):
                for r in d['data_rows']:
                    code = norm(ws.cell(r, 1).value)
                    name = norm(ws.cell(r, 2).value) or code
                    add_row(sname, r, code, name)
                    for col in 'DEFG':
                        summary_cells.add((sname, r, col))
                # 合计行（cat/sum2 的 hj 行，A 中有值才纳入）
                if d['hj']:
                    r = d['hj']
                    code = norm(ws.cell(r, 1).value)
                    name = norm(ws.cell(r, 2).value) or code
                    add_row(sname, r, code, name, is_total=True)
                    for col in 'DEFG':
                        summary_cells.add((sname, r, col))
            elif d['kind'] == 'class':      # 2-分类汇总
                ws2 = self.wa['2-分类汇总']
                for r in range(7, min(ws2.max_row, 67) + 1):
                    a = ws2.cell(r, 1).value
                    b = ws2.cell(r, 2).value
                    if a is None and b is None:
                        continue
                    code = norm(a) or ''
                    name = norm(b) or ''
                    add_row('2-分类汇总', r, code, name)
                    for col in 'CDEF':
                        summary_cells.add(('2-分类汇总', r, col))
            elif d['kind'] == 'top':        # 1-汇总表
                wst = self.wa['1-汇总表']
                for r in range(8, min(wst.max_row, 33) + 1):
                    cname = wst.cell(r, 3).value
                    if not cname:
                        continue
                    add_row('1-汇总表', r, '', norm(cname))
                    for col in 'DEFG':
                        summary_cells.add(('1-汇总表', r, col))
        self.summary_rows = summary_rows
        self.summary_cells = summary_cells
        return summary_rows, summary_cells

    # ---------------- 公式求值（带缓存） ----------------
    def get_val(self, sheet, col, row, depth=0, chain=None):
        key = (sheet, f"{col}{row}")
        if key in self._eval_cache:
            return self._eval_cache[key]
        if depth > 60:
            self._eval_cache[key] = None
            return None
        if chain is None:
            chain = set()
        if key in chain:
            self._eval_cache[key] = None
            return None
        chain = chain | {key}
        v = self.wb[sheet][f"{col}{row}"].value
        if isinstance(v, str) and v.startswith('='):
            ev = self._eval_formula(v, sheet, depth + 1, chain)
            self._eval_cache[key] = ev
            return ev
        self._eval_cache[key] = v
        return v

    def _eval_formula(self, formula, sheet, depth, chain):
        s = formula[1:].strip()
        m = re.match(r'IF\((.*)\)$', s, re.S)
        if m:
            inner = m.group(1)
            parts, d2, cur = [], 0, ''
            for ch in inner:
                if ch == '(':
                    d2 += 1
                elif ch == ')':
                    d2 -= 1
                if ch == ',' and d2 == 0:
                    parts.append(cur)
                    cur = ''
                else:
                    cur += ch
            parts.append(cur)
            if len(parts) == 3:
                cond = self._eval_expr(parts[0], sheet, depth, chain)
                if cond is None:
                    return None
                branch = parts[1] if cond else parts[2]
                return self._eval_expr(branch, sheet, depth, chain)
        return self._eval_expr(s, sheet, depth, chain)

    @staticmethod
    def _to_py(v):
        """Excel 单元格值 → Python 字面量：空白单元格=0；空文本=''；数值=repr；其余文本=repr"""
        if v is None:
            return '0'
        if isinstance(v, str):
            return "''" if v == '' else repr(v)
        return repr(v)

    def _eval_expr(self, expr, sheet, depth, chain):
        s = expr
        # 1) 范围求和函数（模板自带公式：SUM(D7:D26) / SUBTOTAL(9,D7:D9) / ROUND(SUM(F6:F32),2)）
        #    Excel 语义：SUM 包含嵌套 SUBTOTAL 值；SUBTOTAL 忽略嵌套 SUBTOTAL（避免小计重复计算）
        def repl_sum_arg(args, ignore_subtotal):
            r = self._range_sum(args, sheet, depth, chain, ignore_subtotal=ignore_subtotal)
            return 'None' if r is None else repr(r)
        s = re.sub(r'(?i)\bSUM\(([^)]*)\)', lambda m: repl_sum_arg(m.group(1), False), s)
        s = re.sub(r'(?i)\bSUBTOTAL\(\s*\d+\s*,\s*([^)]*)\)', lambda m: repl_sum_arg(m.group(1), True), s)
        # 2) ROUND(x, n) → round(x, n)
        s = re.sub(r'(?i)\bROUND\(([^,]+),\s*(\d+)\)', lambda m: f'round({m.group(1)}, {m.group(2)})', s)

        def repl_cross(m):
            ws, c, r = m.group(1), m.group(2).replace('$', ''), m.group(3).replace('$', '')
            if ws not in self.wb.sheetnames:
                return 'None'
            v = self.get_val(ws, c, int(r), depth, chain)
            return self._to_py(v)
        s = CELL_REF.sub(repl_cross, s)

        def repl_local(m):
            c, r = m.group(1).replace('$', ''), m.group(2).replace('$', '')
            v = self.get_val(sheet, c, int(r), depth, chain)
            return self._to_py(v)
        s = LOCAL_REF.sub(repl_local, s)
        s = re.sub(r'ABS\(([^)]*)\)', r'abs(\1)', s)
        s = re.sub(r'(?<![=<>!])=(?!=)', '==', s)
        try:
            return eval(s, {'__builtins__': {}}, {'abs': abs, 'round': round})
        except Exception:
            return None

    def _range_sum(self, args, sheet, depth, chain, ignore_subtotal=False):
        """SUM/SUBTOTAL 参数求和：支持本表范围 D7:D26、单格 D7、逗号分隔多段。
        按 Excel 语义：文本/空白被忽略；SUBTOTAL 调用时忽略范围内嵌套 SUBTOTAL（防小计重复计算）"""
        total = 0.0
        for part in args.split(','):
            part = part.strip()
            m = re.fullmatch(r'(\$?[A-Z]{1,2})(\$?\d+):(\$?[A-Z]{1,2})(\$?\d+)', part)
            if m:
                c1, r1 = m.group(1).replace('$', ''), int(m.group(2).replace('$', ''))
                c2, r2 = m.group(3).replace('$', ''), int(m.group(4).replace('$', ''))
                col1, col2 = CI(c1), CI(c2)
                for rr in range(min(r1, r2), max(r1, r2) + 1):
                    for cc in range(min(col1, col2), max(col1, col2) + 1):
                        if ignore_subtotal and sheet in self.wb.sheetnames:
                            fv = self.wb[sheet][f"{CL(cc)}{rr}"].value
                            if isinstance(fv, str) and fv.lstrip('=').upper().startswith('SUBTOTAL'):
                                continue   # 嵌套 SUBTOTAL：Excel 忽略其值（小计已含于明细）
                        n = to_num(self.get_val(sheet, CL(cc), rr, depth, chain))
                        if n is not None:
                            total += n
            else:
                m2 = re.fullmatch(r'(\$?[A-Z]{1,2})(\$?\d+)', part)
                if m2:
                    n = to_num(self.get_val(sheet, m2.group(1).replace('$', ''),
                                            int(m2.group(2).replace('$', '')), depth, chain))
                    if n is not None:
                        total += n
                else:
                    return None
        return total

    # ---------------- Step 1 单元格完整性 ----------------
    def check_integrity(self):
        issues = []
        # sheet 列表一致性
        if set(self.wa.sheetnames) != set(self.wb.sheetnames):
            only_a = set(self.wa.sheetnames) - set(self.wb.sheetnames)
            only_b = set(self.wb.sheetnames) - set(self.wa.sheetnames)
            issues.append(('STEP1-结构', '', '', '', '',
                           f'sheet 列表不一致: A独有{only_a} B独有{only_b}'))
        for sname in self.wa.sheetnames:
            if sname not in self.wb.sheetnames:
                continue
            wa, wb = self.wa[sname], self.wb[sname]
            maxr = max(wa.max_row, wb.max_row)
            maxc = max(wa.max_column, wb.max_column)
            for r in range(1, maxr + 1):
                for c in range(1, maxc + 1):
                    col = CL(c)
                    if (sname, r, col) in self.summary_cells:
                        continue   # 汇总单元格由 Step 3 处理
                    av = wa.cell(r, c).value
                    bv = wb.cell(r, c).value
                    if av is None and bv is None:
                        continue
                    if self._same_value(av, bv):
                        continue
                    # 非汇总行内容被改动
                    issues.append(('STEP1-完整性', sname, r, col,
                                   f'A={av!r}', f'B={bv!r}'))
        return issues

    def _same_value(self, a, b):
        """非汇总单元格 A vs B 一致性（数值按容差、文本/其他类型按相等）"""
        an, bn = to_num(a), to_num(b)
        if an is not None and bn is not None:
            return abs(an - bn) <= 1e-9   # 浮点噪音容差
        if isinstance(a, str) or isinstance(b, str):
            return norm(a) == norm(b)
        # 其他类型（datetime/bool/None 等）：Python 相等
        try:
            return bool(a == b) or (a is None and b is None)
        except Exception:
            return a is b

    # ---------------- Step 2 公式合法性 ----------------
    def check_formula(self):
        issues = []
        n_ok_formula = 0
        n_old_formula = 0
        # ① 应处理汇总行必须生成公式（A 有静态值 → B 应为公式）
        for (sname, r, col) in sorted(self.summary_cells):
            av = self.wa[sname].cell(r, CI(col)).value
            bv = self.wb[sname].cell(r, CI(col)).value
            b_is_formula = isinstance(bv, str) and bv.startswith('=')
            if isinstance(av, str) and av.startswith('='):
                # A 中已是旧公式（如 SUBTOTAL/勾稽校验）→ B 保留即合规
                if b_is_formula:
                    n_old_formula += 1
                else:
                    issues.append(('STEP2-旧公式丢失', sname, r, col,
                                   f'A旧公式={av}', f'B={bv!r}'))
                continue
            if av in (None, ''):
                # 空汇总行：Skill 策略留空 → B 无公式合规；B 有公式则异常
                if b_is_formula:
                    issues.append(('STEP2-空行新增公式', sname, r, col,
                                   'A为空', f'B={bv}'))
                continue
            # A 有静态值 → B 必须生成公式
            if b_is_formula:
                n_ok_formula += 1
                # ③ 公式错误文本检测
                for mk in ERR_MARKERS:
                    if mk in bv.upper():
                        issues.append(('STEP2-公式错误', sname, r, col,
                                       f'公式含{mk}', bv))
                        break
                # 公式求值错误检测（计算失败且引用非占位符）
                val = self.get_val(sname, col, r)
                if val is None and not placeholder_code(av):
                    src = self.wav[sname].cell(r, CI(col)).value
                    issues.append(('STEP2-公式计算失败', sname, r, col,
                                   f'A基准={src!r}', f'B公式={bv}'))
            else:
                issues.append(('STEP2-缺公式', sname, r, col,
                               f'A静态值={av}', f'B={bv!r}'))
        # ② 非汇总行不能新增公式
        for sname in self.wb.sheetnames:
            wb = self.wb[sname]
            for r in range(1, wb.max_row + 1):
                for c in range(1, wb.max_column + 1):
                    col = CL(c)
                    if (sname, r, col) in self.summary_cells:
                        continue
                    bv = wb.cell(r, c).value
                    if isinstance(bv, str) and bv.startswith('='):
                        av = self.wa[sname].cell(r, c).value
                        if isinstance(av, str) and av.startswith('='):
                            continue   # A 中已有旧公式（非汇总行旧公式，如叶子表内部公式）
                        issues.append(('STEP2-非汇总行新增公式', sname, r, col,
                                       f'A={av!r}', f'B={bv}'))
        return issues, n_ok_formula, n_old_formula

    # ---------------- Step 3 数值一致性 ----------------
    def check_numeric(self):
        issues = []
        n_checked = 0
        n_ok = 0
        for (sname, r, col) in sorted(self.summary_cells):
            av = self.wav[sname].cell(r, CI(col)).value   # A 静态值（基准）
            bv = self.wb[sname].cell(r, CI(col)).value    # B 公式
            if not (isinstance(bv, str) and bv.startswith('=')):
                continue   # 无公式的行（空行/旧公式）由 Step 2 处理
            pc = placeholder_code(av)
            if pc:
                issues.append(('STEP3-源异常占位符', sname, r, col,
                               f'A={av}', f'B公式={bv}'))
                continue
            calc = self.get_val(sname, col, r)
            a_num = to_num(av)
            if calc is None:
                # 计算失败：源占位符或公式错误（Step 2 已报计算失败，这里补充数值比对不可行）
                continue
            if isinstance(calc, str):
                # IF 分支返回 ""（账面为0时增值率留空）：与 A 基准 0/空 语义等价 → 一致
                if calc == '' and (a_num == 0 or a_num is None):
                    n_checked += 1
                    n_ok += 1
                    continue
                issues.append(('STEP3-计算值非数值', sname, r, col,
                               f'A基准={av!r}', f'B计算={calc!r}'))
                continue
            if a_num is None:
                if av not in (None, ''):
                    issues.append(('STEP3-基准值不可解析', sname, r, col,
                                   f'A={av!r}', f'B计算={calc}'))
                continue
            n_checked += 1
            diff = abs(calc - a_num)
            if diff <= AMOUNT_TOL:
                n_ok += 1
            else:
                issues.append(('STEP3-数值不一致', sname, r, col,
                               f'A基准={av}', f'B计算={calc}', f'|差|={round(diff,4)}',
                               f'公式={bv}'))
        return issues, n_checked, n_ok

    # ---------------- 汇总 ----------------
    def run(self):
        self.identify()
        issues_step1 = self.check_integrity()
        # 明细行一致性：Step 1 中非汇总行若发现改动 → 比对无效终止
        # （非汇总行包含表头/明细/备注/空行；汇总行由 Step 3 单独比对）
        if issues_step1:
            return {'valid': False, 'step1': issues_step1, 'step2': [], 'step3': [],
                    'reason': '明细行内容不一致，比对无效', 'stats': None}
        issues_step2, n_ok_formula, n_old_formula = self.check_formula()
        issues_step3, n_checked, n_ok = self.check_numeric()
        all_issues = issues_step2 + issues_step3
        summary_rows_count = len(self.summary_rows)
        return {
            'valid': not all_issues,
            'step1': issues_step1,
            'step2': issues_step2,
            'step3': issues_step3,
            'stats': {
                'summary_rows': summary_rows_count,
                'formula_ok': n_ok_formula,
                'old_formula': n_old_formula,
                'numeric_checked': n_checked,
                'numeric_ok': n_ok,
            },
        }


# ---------------- 报告输出 ----------------
def write_report(result, out_dir, name_a, name_b):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '校对结论'
    st = result['stats']
    if result['valid']:
        ws.append(['评估明细表公式恢复 自动校对报告'])
        ws.append(['A基准文件', name_a])
        ws.append(['B输出文件', name_b])
        ws.append(['结论', '✅ 自动校验全部通过，无需人工校对'])
        ws.append([])
        ws.append(['汇总总行数', st['summary_rows']])
        ws.append(['成功生成公式数量', st['formula_ok']])
        ws.append(['保留旧公式数量', st['old_formula']])
        ws.append(['数值一致性比对格数', st['numeric_checked']])
        ws.append(['数值一致格数', st['numeric_ok']])
        ws.append(['数值不一致', st['numeric_checked'] - st['numeric_ok']])
    else:
        ws.append(['评估明细表公式恢复 自动校对报告'])
        ws.append(['A基准文件', name_a])
        ws.append(['B输出文件', name_b])
        if not result['valid'] and result.get('reason'):
            ws.append(['结论', '比对无效：' + result['reason']])
        else:
            ws.append(['结论', f'存在 {len(result["step2"]) + len(result["step3"])} 项异常，仅清单内项目需人工复核'])
    for i, w in enumerate((24, 12, 10, 10, 60, 60, 20, 60), 1):
        ws.column_dimensions[CL(i)].width = w

    ws2 = wb.create_sheet('异常清单')
    ws2.append(['异常类型', 'sheet', '行号', '列', '科目编码/字段', 'A基准', 'B输出', '说明'])
    n = 0
    for grp in ('step1', 'step2', 'step3'):
        for it in result[grp]:
            n += 1
            ws2.append([it[0], it[1], it[2], it[3], it[4] if len(it) > 4 else '',
                        it[5] if len(it) > 5 else '', it[6] if len(it) > 6 else '',
                        '；'.join(it[7:]) if len(it) > 7 else ''])
    ws2.append([])
    ws2.append(['异常总数', n])
    for i, w in enumerate((22, 26, 8, 8, 16, 34, 34, 60), 1):
        ws2.column_dimensions[CL(i)].width = w
    ws2.freeze_panes = 'A2'
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, '校对报告.xlsx')
    wb.save(out)
    return out


def main():
    if len(sys.argv) < 3:
        print('用法: python audit_restore.py <A基准文件.xlsx> <B输出文件.xlsx> [输出目录]')
        sys.exit(1)
    path_a, path_b = sys.argv[1], sys.argv[2]
    out_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.dirname(os.path.abspath(path_b))
    aud = Auditor(path_a, path_b)
    result = aud.run()
    st = result['stats']
    print(f"A={os.path.basename(path_a)}")
    print(f"B={os.path.basename(path_b)}")
    if not result['valid'] and result.get('reason'):
        print(f"结论: 比对无效 - {result['reason']}")
        print(f"  STEP1 异常 {len(result['step1'])} 项")
        for it in result['step1'][:10]:
            print(f"    [{it[0]}] {it[1]}!{it[2]}{it[3]}: {it[4]} vs {it[5]}")
        return
    print(f"STEP1 完整性: 通过（非汇总行无改动）")
    print(f"STEP2 公式合法性: 缺公式/新增/错误 {len(result['step2'])} 项（成功公式 {st['formula_ok']}，旧公式保留 {st['old_formula']}）")
    print(f"STEP3 数值一致性: 比对 {st['numeric_checked']} 格，一致 {st['numeric_ok']}，不一致 {st['numeric_checked'] - st['numeric_ok']}")
    total = len(result['step2']) + len(result['step3'])
    if total == 0:
        print(f"结论: ✅ 自动校验全部通过，无需人工校对（汇总总行数 {st['summary_rows']}，成功生成公式 {st['formula_ok']}）")
    else:
        print(f"结论: ⚠️ 存在 {total} 项异常，仅清单内项目需人工复核")
        for grp in ('step2', 'step3'):
            for it in result[grp]:
                print(f"    [{it[0]}] {it[1]}!R{it[2]}{it[3]}: A={it[4]} | B={it[5]} | {('；'.join(it[7:]) if len(it) > 7 else it[6] if len(it) > 6 else '')}")
    rep = write_report(result, out_dir, os.path.basename(path_a), os.path.basename(path_b))
    print('报告:', rep)


if __name__ == '__main__':
    main()
