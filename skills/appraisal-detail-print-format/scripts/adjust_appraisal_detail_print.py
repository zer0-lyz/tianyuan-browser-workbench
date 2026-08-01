#!/usr/bin/env python3
"""Format exported appraisal detail workbooks for clean printing."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


FIELD_MARKER = re.compile(r"^(index|field\d+|bookSumValue|evaluateSumValue|difference|growRate)$", re.I)
RMB_VALUES = {"人民币", "人民幣", "RMB", "CNY"}
A4_PORTRAIT_WIDTH_POINTS = 595.276
A4_PORTRAIT_HEIGHT_POINTS = 841.89
DEFAULT_PAGE_HEIGHT_FACTOR = 0.95
DEFAULT_MIN_COMPRESS_ROW_HEIGHT = 11.2


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def is_blankish(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return True
        if text.startswith("="):
            return False
        try:
            return float(text.replace(",", "").replace("%", "")) == 0
        except ValueError:
            return False
    if isinstance(value, (int, float)):
        return value == 0
    return False


def is_detail_seq(value) -> bool:
    if isinstance(value, (int, float)):
        return float(value).is_integer() and value > 0
    return bool(re.fullmatch(r"\d+", norm(value)))


def meaningful(value) -> bool:
    text = norm(value)
    return bool(text) and not FIELD_MARKER.match(text)


def area_bounds(ws):
    if ws.print_area:
        ref = str(ws.print_area).split("!")[-1].replace("'", "").replace("$", "")
        return range_boundaries(ref)
    return 1, 1, ws.max_column, ws.max_row


def find_header(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        headers = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        book_cols = [col for col, head in headers.items() if head == "账面价值"]
        if book_cols:
            return row, headers, book_cols[0]
    return None, {}, None


def detail_rows(ws, header_row):
    return [row for row in range(header_row + 1, ws.max_row + 1) if is_detail_seq(ws.cell(row, 1).value)]


def find_last_print_row(ws, max_col):
    last = 1
    for row in range(1, ws.max_row + 1):
        text = " ".join(str(ws.cell(row, col).value) for col in range(1, max_col + 1) if ws.cell(row, col).value is not None)
        vals = [ws.cell(row, col).value for col in range(1, max_col + 1)]
        if not any(meaningful(v) for v in vals):
            continue
        if all(FIELD_MARKER.match(norm(v)) for v in vals if norm(v)):
            continue
        last = row
        if "评估机构" in text:
            last = row
    return last


def cell_has_border(cell):
    border = cell.border
    return any(getattr(side, "style", None) for side in (border.left, border.right, border.top, border.bottom))


def find_table_bounds(ws):
    max_row_scan = min(ws.max_row, 80)
    max_col_scan = ws.max_column
    last_row = 1
    last_col = 1
    for row in range(1, max_row_scan + 1):
        for col in range(1, max_col_scan + 1):
            cell = ws.cell(row, col)
            if meaningful(cell.value) or cell_has_border(cell):
                last_row = max(last_row, row)
                last_col = max(last_col, col)

    # Avoid styled blank columns to the right of the real header/table.
    header_candidates = []
    for row in range(1, min(ws.max_row, 20) + 1):
        row_last = 0
        for col in range(1, max_col_scan + 1):
            cell = ws.cell(row, col)
            if meaningful(cell.value) or cell_has_border(cell):
                row_last = col
        if row_last:
            header_candidates.append(row_last)
    if header_candidates:
        last_col = min(last_col, max(header_candidates))
    return 1, last_col


def visible_cols(ws):
    min_col, _, max_col, _ = area_bounds(ws)
    return [col for col in range(min_col, max_col + 1) if not ws.column_dimensions[get_column_letter(col)].hidden]


def width_sum(ws, cols):
    return sum(ws.column_dimensions[get_column_letter(col)].width or 8.43 for col in cols)


def resize_visible_cols(ws, target_width, min_width=3.6, max_width=60.0):
    cols = visible_cols(ws)
    before = width_sum(ws, cols)
    if not before:
        return 0.0, 0.0

    widths = []
    factor = target_width / before
    for col in cols:
        old = ws.column_dimensions[get_column_letter(col)].width or 8.43
        widths.append((col, max(min_width, min(max_width, old * factor))))

    for _ in range(10):
        current = sum(width for _, width in widths)
        diff = target_width - current
        if abs(diff) < 0.03:
            break
        adjustable = [
            idx
            for idx, (_, width) in enumerate(widths)
            if (diff > 0 and width < max_width - 0.01) or (diff < 0 and width > min_width + 0.01)
        ]
        if not adjustable:
            break
        delta = diff / len(adjustable)
        for idx in adjustable:
            col, width = widths[idx]
            widths[idx] = (col, max(min_width, min(max_width, width + delta)))

    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = round(width, 2)
    return round(before, 1), round(width_sum(ws, cols), 1)


def hide_empty_columns(ws):
    changes = []
    header_row, headers, book_col = find_header(ws)
    if not header_row or not book_col:
        return changes

    rows = detail_rows(ws, header_row)
    if not rows:
        return changes

    protected = {"序号", "编号", "项目编号"}
    for col in range(1, book_col):
        head = headers.get(col, "")
        if not head or head in protected or "名称" in head or "对象" in head:
            continue
        if all(is_blankish(ws.cell(row, col).value) for row in rows):
            letter = get_column_letter(col)
            if not ws.column_dimensions[letter].hidden:
                ws.column_dimensions[letter].hidden = True
                changes.append(f"{letter}:{head}")

    currency_col = next((col for col, head in headers.items() if head == "币种"), None)
    rate_col = next((col for col, head in headers.items() if head == "评估基准日汇率"), None)
    if currency_col and rate_col:
        currencies = [norm(ws.cell(row, currency_col).value) for row in rows if norm(ws.cell(row, currency_col).value)]
        if currencies and all(value in RMB_VALUES for value in currencies):
            letter = get_column_letter(rate_col)
            if not ws.column_dimensions[letter].hidden:
                ws.column_dimensions[letter].hidden = True
                changes.append(f"{letter}:评估基准日汇率")

    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, ws.max_column + 1):
            if "成新率" not in norm(ws.cell(row, col).value):
                continue
            rows2 = detail_rows(ws, row)
            if rows2 and all(is_blankish(ws.cell(r, col).value) for r in rows2):
                letter = get_column_letter(col)
                if not ws.column_dimensions[letter].hidden:
                    ws.column_dimensions[letter].hidden = True
                    changes.append(f"{letter}:{ws.cell(row, col).value}")
    return changes


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("="):
            return None
        try:
            return float(text.replace(",", "").replace("%", ""))
        except ValueError:
            return None
    return None


def is_empty_summary_sheet(ws):
    title = ws.title
    if "汇总" not in title or title.startswith(("1-", "2-")):
        return False
    header_row, headers, _ = find_header(ws)
    if not header_row:
        return False
    amount_cols = [
        col
        for col, head in headers.items()
        if head in {"账面价值", "评估价值", "增值额", "增值率%", "增值率％"}
    ]
    if not amount_cols:
        return False
    nums = []
    for row in range(header_row + 1, ws.max_row + 1):
        for col in amount_cols:
            number = parse_number(ws.cell(row, col).value)
            if number is not None:
                nums.append(number)
    return bool(nums) and all(abs(number) < 1e-9 for number in nums)


def row_height_sum(ws, last_row):
    total = 0.0
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        total += ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
    return total


def paper_height_points(ws):
    if ws.page_setup.orientation == "landscape":
        return A4_PORTRAIT_WIDTH_POINTS
    return A4_PORTRAIT_HEIGHT_POINTS


def safe_page_height_limit(ws, page_height_factor=DEFAULT_PAGE_HEIGHT_FACTOR):
    printable_height = paper_height_points(ws) - (ws.page_margins.top + ws.page_margins.bottom) * 72
    return max(1.0, printable_height * page_height_factor)


def tighten_rows_if_needed(ws, last_row, limit=None, page_height_factor=DEFAULT_PAGE_HEIGHT_FACTOR, min_row_height=DEFAULT_MIN_COMPRESS_ROW_HEIGHT):
    if limit is None:
        limit = safe_page_height_limit(ws, page_height_factor)
    total = row_height_sum(ws, last_row)
    if last_row > 40 or total <= limit:
        return total, total
    factor = limit / total
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        old = ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
        floor = min_row_height if old >= min_row_height else old
        ws.row_dimensions[row].height = min(old, max(floor, round(old * factor, 2)))
    return total, row_height_sum(ws, last_row)


def apply_print_setup(ws, args):
    min_col, max_col = find_table_bounds(ws)
    last_row = find_last_print_row(ws, max_col)
    ws.print_area = f"{get_column_letter(min_col)}1:{get_column_letter(max_col)}{last_row}"
    ws.row_breaks.brk = []
    ws.col_breaks.brk = []
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    # Use OOXML fit-to-page mode so Excel/WPS on Windows prints every visible
    # column on one page wide; the page height remains unrestricted.
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_margins.left = args.margin
    ws.page_margins.right = args.margin
    ws.page_margins.top = args.top_margin
    ws.page_margins.bottom = args.bottom_margin
    ws.page_margins.header = args.header_margin
    ws.page_margins.footer = args.footer_margin
    ws.print_options.horizontalCentered = True
    before_h, after_h = tighten_rows_if_needed(
        ws,
        last_row,
        page_height_factor=getattr(args, "page_height_factor", DEFAULT_PAGE_HEIGHT_FACTOR),
        min_row_height=getattr(args, "min_compress_row_height", DEFAULT_MIN_COMPRESS_ROW_HEIGHT),
    )
    before_w, after_w = resize_visible_cols(ws, args.target_width)
    return last_row, before_w, after_w, before_h, after_h


def parse_args(argv):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to an appraisal detail .xlsx/.xlsm workbook")
    parser.add_argument("--target-width", type=float, default=143.0, help="Total visible column width to fit page")
    parser.add_argument("--margin", type=float, default=0.45, help="Left/right page margins in inches")
    parser.add_argument("--top-margin", type=float, default=0.30)
    parser.add_argument("--bottom-margin", type=float, default=0.60)
    parser.add_argument("--header-margin", type=float, default=0.12)
    parser.add_argument("--footer-margin", type=float, default=0.30)
    parser.add_argument("--page-height-factor", type=float, default=DEFAULT_PAGE_HEIGHT_FACTOR, help="Safety factor for WPS print-preview page height")
    parser.add_argument("--min-compress-row-height", type=float, default=DEFAULT_MIN_COMPRESS_ROW_HEIGHT)
    parser.add_argument("--keep-empty-summary", action="store_true", help="Do not hide no-data summary sheets")
    parser.add_argument("--dry-run", action="store_true", help="Report changes without saving")
    parser.add_argument("--no-backup", action="store_true", help="Do not create timestamped backup")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv or sys.argv[1:])
    path = Path(args.workbook).expanduser()
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    backup = None
    if not args.dry_run and not args.no_backup:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = path.with_name(f"{path.stem}.打印格式调整前备份-{stamp}{path.suffix}")
        copy2(path, backup)

    wb = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    report = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if not args.keep_empty_summary and is_empty_summary_sheet(ws):
            ws.sheet_state = "hidden"
            report.append((ws.title, ["sheet:hidden-empty-summary"], (0, 0, 0, 0, 0)))
            continue
        hidden = hide_empty_columns(ws)
        metrics = apply_print_setup(ws, args)
        report.append((ws.title, hidden, metrics))

    if not args.dry_run:
        wb.save(path)

    if backup:
        print(f"backup={backup}")
    print(f"visible_sheets={len(report)} dry_run={args.dry_run}")
    for title, hidden, (last_row, before_w, after_w, before_h, after_h) in report:
        hidden_text = ",".join(hidden) if hidden else "-"
        print(
            f"{title}\tlast_row={last_row}\twidth={before_w}->{after_w}"
            f"\theight={round(before_h,1)}->{round(after_h,1)}\thidden={hidden_text}"
        )


if __name__ == "__main__":
    main()
