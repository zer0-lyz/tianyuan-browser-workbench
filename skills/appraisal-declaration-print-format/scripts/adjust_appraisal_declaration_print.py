#!/usr/bin/env python3
"""Convert an appraisal detail workbook into a declaration print workbook."""

from __future__ import annotations

import argparse
import importlib.util
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from copy import copy
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


BASE_SCRIPT = Path(
    os.environ.get(
        "TIANYUAN_DETAIL_PRINT_SCRIPT",
        Path(__file__).resolve().parents[2]
        / "appraisal-detail-print-format"
        / "scripts"
        / "adjust_appraisal_detail_print.py",
    )
)
RESULT_HEADERS = ("评估价值", "增值额", "增值率")


def load_base_module():
    if not BASE_SCRIPT.exists():
        raise SystemExit(f"Required base skill script not found: {BASE_SCRIPT}")
    spec = importlib.util.spec_from_file_location("appraisal_detail_print_base", BASE_SCRIPT)
    if spec is None or spec.loader is None:
        raise SystemExit(f"Cannot load base skill script: {BASE_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()



def base_is_detail_seq(value) -> bool:
    if isinstance(value, (int, float)):
        return float(value).is_integer() and value > 0
    return bool(re.fullmatch(r"\d+", norm(value)))


def is_result_header(value) -> bool:
    text = norm(value)
    return any(text == header or text.startswith(header + "%") or text.startswith(header + "％") for header in RESULT_HEADERS)


def merged_column_span(ws, row: int, col: int):
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_col, merged.max_col
    return col, col


def hide_result_columns(ws):
    hidden = set()
    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, ws.max_column + 1):
            if not is_result_header(ws.cell(row, col).value):
                continue
            min_col, max_col = merged_column_span(ws, row, col)
            for target_col in range(min_col, max_col + 1):
                letter = get_column_letter(target_col)
                ws.column_dimensions[letter].hidden = True
                hidden.add(letter)
    return sorted(hidden)


def replace_title(ws):
    replacements = []
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell.value, str) and "评估明细表" in cell.value:
                old = cell.value
                cell.value = old.replace("评估明细表", "评估申报明细表")
                replacements.append(cell.coordinate)
    return replacements



def is_blank_display_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return True
        if text.startswith("="):
            return True
        try:
            return float(text.replace(",", "").replace("%", "")) == 0
        except ValueError:
            return False
    if isinstance(value, (int, float)):
        return value == 0
    return False


def is_formula_text(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def has_substantive_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return False
        if text.startswith("="):
            return True
        return True
    return True


def row_text(ws, row: int, max_col: int) -> str:
    return "".join(str(ws.cell(row, col).value or "") for col in range(1, max_col + 1))


def compact_row_text(ws, row: int, max_col: int) -> str:
    return re.sub(r"\s+", "", row_text(ws, row, max_col))


def is_first_total_row(ws, row: int, max_col: int) -> bool:
    return "合计" in compact_row_text(ws, row, max_col)


def is_summary_row(ws, row: int, max_col: int) -> bool:
    text = compact_row_text(ws, row, max_col)
    return any(token in text for token in ("合计", "减：", "减:", "减值准备", "预计损失", "坏账准备"))


def find_table_header_row(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        text = "".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1))
        if "序号" in text and any(token in text for token in ("账面价值", "备注", "名称", "开户银行")):
            return row
    return None


def effective_table_bounds(ws):
    max_col = 1
    max_row = min(ws.max_row, 120)
    for row in range(1, max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if has_substantive_value(cell.value) or any(getattr(side, "style", None) for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)):
                max_col = max(max_col, col)
    return max_col


def find_first_total_row(ws, header_row: int, max_col: int):
    for row in range(header_row + 1, ws.max_row + 1):
        if is_first_total_row(ws, row, max_col):
            return row
    return None


def header_labels(ws, header_row: int):
    labels = {}
    for col in range(1, ws.max_column + 1):
        labels[col] = "".join(norm(ws.cell(row, col).value) for row in range(max(1, header_row - 1), min(ws.max_row, header_row + 1) + 1))
    return labels


def detail_key_columns(ws, header_row: int):
    labels = header_labels(ws, header_row)
    object_cols = [
        col for col, label in labels.items()
        if any(token in label for token in ("名称", "对象", "开户银行", "账号", "内容", "项目", "科目", "资产"))
    ]
    amount_cols = [
        col for col, label in labels.items()
        if any(token in label for token in ("账面价值", "金额", "净值", "成本", "余额"))
    ]
    return object_cols, amount_cols


def cell_is_real_data(value) -> bool:
    if is_blank_display_value(value) or is_formula_text(value):
        return False
    return has_substantive_value(value)


def row_has_object_or_amount(ws, row: int, object_cols, amount_cols) -> bool:
    for col in object_cols + amount_cols:
        if ws.column_dimensions[get_column_letter(col)].hidden:
            continue
        if cell_is_real_data(ws.cell(row, col).value):
            return True
    return False


def clear_fake_detail_rows(ws):
    header_row = find_table_header_row(ws)
    if not header_row:
        return 0, None, None, None
    max_col = effective_table_bounds(ws)
    first_total = find_first_total_row(ws, header_row, max_col)
    if not first_total:
        return 0, header_row, None, max_col
    object_cols, amount_cols = detail_key_columns(ws, header_row)
    cleared = 0
    for row in range(header_row + 1, first_total):
        if not base_is_detail_seq(ws.cell(row, 1).value):
            continue
        if row_has_object_or_amount(ws, row, object_cols, amount_cols):
            continue
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
        cleared += 1
    return cleared, header_row, first_total, max_col


def find_last_data_row(ws, header_row: int, first_total: int):
    object_cols, amount_cols = detail_key_columns(ws, header_row)
    last = header_row
    for row in range(header_row + 1, first_total):
        if row_has_object_or_amount(ws, row, object_cols, amount_cols):
            last = row
    return last


def sanitize_and_classify_detail_rows(ws):
    """Clear template-only sequence rows and report whether real detail remains.

    This intentionally never inserts or deletes rows. Row count changes require
    a reviewed WPS-native operation because print pagination is host-specific.
    """
    header_row = find_table_header_row(ws)
    if not header_row:
        return 0, False, None, None
    max_col = effective_table_bounds(ws)
    first_total = find_first_total_row(ws, header_row, max_col)
    if not first_total:
        return 0, False, header_row, None
    object_cols, amount_cols = detail_key_columns(ws, header_row)
    cleared = 0
    has_real_detail = False
    for row in range(header_row + 1, first_total):
        has_data = row_has_object_or_amount(ws, row, object_cols, amount_cols)
        if has_data:
            has_real_detail = True
            continue
        if not base_is_detail_seq(ws.cell(row, 1).value):
            continue
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
        cleared += 1
    return cleared, has_real_detail, header_row, first_total


CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")
SUM_RANGE_RE = re.compile(r"SUM\((\$?[A-Z]{1,3})(\$?)(\d+):(\$?[A-Z]{1,3})(\$?)(\d+)\)")


def shift_formula_for_insert(formula: str, insert_at: int, count: int, origin_row: int) -> str:
    saved_ranges = []

    def save_sum(match):
        c1, a1, r1, c2, a2, r2 = match.groups()
        start = int(r1)
        end = int(r2)
        if start >= insert_at:
            start += count
        if end >= insert_at or (end == insert_at - 1 and origin_row >= insert_at):
            end += count
        token = f"__SUMRANGE{len(saved_ranges)}__"
        saved_ranges.append((token, f"SUM({c1}{a1}{start}:{c2}{a2}{end})"))
        return token

    protected = SUM_RANGE_RE.sub(save_sum, formula)

    def ref_repl(match):
        col, absolute, row_text = match.groups()
        row = int(row_text)
        if row >= insert_at:
            row += count
        return f"{col}{absolute}{row}"

    updated = CELL_REF_RE.sub(ref_repl, protected)
    for token, value in saved_ranges:
        updated = updated.replace(token, value)
    return updated


def copy_row_format(ws, source_row: int, target_row: int, max_col: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)
        dst.value = None


def row_height_sum_until(ws, last_row: int) -> float:
    total = 0.0
    for row in range(1, last_row + 1):
        if not ws.row_dimensions[row].hidden:
            total += ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
    return total


def insert_formula_safe_blank_rows(ws, insert_at: int, count: int, template_row: int, max_col: int):
    if count <= 0:
        return 0
    formulas = []
    for row in range(insert_at, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if is_formula_text(value):
                formulas.append((row, col, value))
    ws.insert_rows(insert_at, count)
    for offset in range(count):
        copy_row_format(ws, template_row, insert_at + offset, max_col)
    for old_row, col, formula in formulas:
        new_row = old_row + count
        ws.cell(new_row, col).value = shift_formula_for_insert(formula, insert_at, count, old_row)
    return count


def fill_short_table_to_page(ws, target_height: float = 500.0, max_insert: int = 8):
    cleared, header_row, first_total, max_col = clear_fake_detail_rows(ws)
    if not header_row or not first_total:
        return cleared, 0
    last_data = find_last_data_row(ws, header_row, first_total)
    # Preserve the complete summary block. New rows belong immediately after the last real detail row.
    insert_at = last_data + 1
    current_height = row_height_sum_until(ws, first_total)
    template_row = insert_at if insert_at < first_total else last_data
    row_height = ws.row_dimensions[template_row].height or ws.sheet_format.defaultRowHeight or 15
    need = min(max_insert, max(0, int((target_height - current_height) // max(row_height, 1))))
    inserted = insert_formula_safe_blank_rows(ws, insert_at, need, template_row, max_col)
    return cleared, inserted


DEFAULT_TABLE_ROW_HEIGHT = 16.5
MIN_COMPRESSED_ROW_HEIGHT = 14.75
MAX_AUTO_INSERT_ROWS = 80
LONG_TABLE_REAL_ROWS = 20
FOOTER_GAP_ROWS = 1
SHORT_TABLE_PAGE_HEIGHT_FACTOR = 0.95
LONG_TABLE_PAGE_HEIGHT_FACTOR = 0.93
UNDERFILLED_LAST_PAGE_REAL_ROWS = 8
UNDERFILLED_LAST_PAGE_BLANK_RATIO = 0.35


def visible_row_height(ws, row: int, fallback: float = DEFAULT_TABLE_ROW_HEIGHT) -> float:
    return ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or fallback


def normalize_table_row_heights(ws, header_row: int, summary_end: int, row_height: float):
    changed = 0
    for row in range(header_row, summary_end + 1):
        if ws.row_dimensions[row].hidden:
            continue
        if abs(visible_row_height(ws, row) - row_height) > 0.01:
            ws.row_dimensions[row].height = row_height
            changed += 1
    return changed


def print_area_bounds(ws):
    if not ws.print_area:
        return 1, 1, ws.max_column, ws.max_row
    ref = str(ws.print_area).split()[0].split("!", 1)[-1].replace("$", "").replace("'", "")
    return range_boundaries(ref)


def excel_col_width_to_points(width):
    width = width or 8.43
    pixels = width * 12 if width < 1 else width * 7 + 5
    return pixels * 0.75


def effective_page_height_points(ws) -> float:
    paper_width = 841.89
    paper_height = 595.276
    printable_width = paper_width - (ws.page_margins.left + ws.page_margins.right) * 72
    printable_height = paper_height - (ws.page_margins.top + ws.page_margins.bottom) * 72
    min_col, _, max_col, _ = print_area_bounds(ws)
    sheet_width = 0.0
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        if not ws.column_dimensions[letter].hidden:
            sheet_width += excel_col_width_to_points(ws.column_dimensions[letter].width)
    scale = min(1.0, printable_width / sheet_width) if sheet_width else 1.0
    return printable_height / max(scale, 0.01)


def repeated_header_height(ws, body_start: int) -> float:
    match = re.search(r"\$(\d+):\$(\d+)", str(ws.print_title_rows or ""))
    if match:
        lo, hi = map(int, match.groups())
        return sum(visible_row_height(ws, row) for row in range(lo, hi + 1) if not ws.row_dimensions[row].hidden)
    return sum(visible_row_height(ws, row) for row in range(1, body_start) if not ws.row_dimensions[row].hidden)


def paginate_table_rows(ws, body_rows, body_start: int, page_height_factor: float = 1.0):
    cap = effective_page_height_points(ws) * page_height_factor
    first_header = sum(visible_row_height(ws, row) for row in range(1, body_start) if not ws.row_dimensions[row].hidden)
    repeat_header = repeated_header_height(ws, body_start)
    used = first_header
    pages = [[]]
    for row in body_rows:
        row_h = visible_row_height(ws, row) if isinstance(row, int) else DEFAULT_TABLE_ROW_HEIGHT
        if used + row_h > cap + 1e-6:
            pages.append([row])
            used = repeat_header + row_h
        else:
            pages[-1].append(row)
            used += row_h
    return pages, used, cap


def find_summary_end(ws, first_total: int, max_col: int):
    last = first_total
    for row in range(first_total + 1, ws.max_row + 1):
        if is_summary_row(ws, row, max_col):
            last = row
        elif row > last + 1:
            break
    return last


def detail_real_rows(ws, header_row: int, first_total: int):
    object_cols, amount_cols = detail_key_columns(ws, header_row)
    rows = []
    for row in range(header_row + 1, first_total):
        if ws.row_dimensions[row].hidden:
            continue
        if row_has_object_or_amount(ws, row, object_cols, amount_cols):
            rows.append(row)
    return rows


def is_blank_template_row(ws, row: int, max_col: int) -> bool:
    for col in range(1, max_col + 1):
        if cell_is_real_data(ws.cell(row, col).value):
            return False
    return True


def shift_formula_for_row_delete(formula: str, deleted_row: int) -> str:
    saved_ranges = []

    def save_sum(match):
        c1, a1, r1, c2, a2, r2 = match.groups()
        start = int(r1)
        end = int(r2)
        if start > deleted_row:
            start -= 1
        if end >= deleted_row:
            end -= 1
        token = f"__SUMRANGE{len(saved_ranges)}__"
        saved_ranges.append((token, f"SUM({c1}{a1}{start}:{c2}{a2}{max(start, end)})"))
        return token

    protected = SUM_RANGE_RE.sub(save_sum, formula)

    def ref_repl(match):
        col, absolute, row_text = match.groups()
        row = int(row_text)
        if row > deleted_row:
            row = max(1, row - 1)
        return f"{col}{absolute}{row}"

    updated = CELL_REF_RE.sub(ref_repl, protected)
    for token, value in saved_ranges:
        updated = updated.replace(token, value)
    return updated


def update_print_area_for_row_delta(ws, row_at: int, delta: int):
    if not ws.print_area:
        return
    parts = []
    for area in str(ws.print_area).split():
        if "!" not in area:
            continue
        title, coords = area.split("!", 1)
        min_col, min_row, max_col, max_row = range_boundaries(coords.replace("$", "").replace("'", ""))
        if max_row >= row_at:
            max_row += delta
        parts.append(f"{title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}")
    if parts:
        ws.print_area = " ".join(parts)


def delete_one_row_preserving_formulas(ws, row_to_delete: int):
    formulas = [
        (cell.row, cell.column, cell.value)
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=") and cell.row != row_to_delete
    ]
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    ws.merged_cells.ranges = set()
    ws.delete_rows(row_to_delete, 1)
    for old_row, col, formula in formulas:
        target_row = old_row - 1 if old_row > row_to_delete else old_row
        ws.cell(target_row, col).value = shift_formula_for_row_delete(formula, row_to_delete)
    for item in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(item)
        if max_row < row_to_delete:
            pass
        elif min_row > row_to_delete:
            min_row -= 1
            max_row -= 1
        elif min_row <= row_to_delete <= max_row:
            max_row -= 1
        if min_row <= max_row:
            ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    update_print_area_for_row_delta(ws, row_to_delete, -1)


def insert_rows_preserving_formulas(ws, insert_at: int, count: int, template_row: int, max_col: int, row_height: float):
    if count <= 0:
        return 0
    styles = []
    for col in range(1, ws.max_column + 1):
        src = ws.cell(template_row, col)
        styles.append((copy(src._style), copy(src.font), copy(src.fill), copy(src.border), copy(src.alignment), src.number_format, copy(src.protection)))
    formulas = [
        (cell.row, cell.column, cell.value)
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    ws.merged_cells.ranges = set()
    ws.insert_rows(insert_at, count)
    for row in range(insert_at, insert_at + count):
        ws.row_dimensions[row].height = row_height
        for col, values in enumerate(styles, 1):
            dst = ws.cell(row, col)
            dst._style = copy(values[0])
            dst.font = copy(values[1])
            dst.fill = copy(values[2])
            dst.border = copy(values[3])
            dst.alignment = copy(values[4])
            dst.number_format = values[5]
            dst.protection = copy(values[6])
            dst.value = None
    for old_row, col, formula in formulas:
        target_row = old_row + count if old_row >= insert_at else old_row
        ws.cell(target_row, col).value = shift_formula_for_insert(formula, insert_at, count, old_row)
    for item in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(item)
        if min_row >= insert_at:
            min_row += count
            max_row += count
        elif max_row >= insert_at:
            max_row += count
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    update_print_area_for_row_delta(ws, insert_at, count)
    return count


def compress_until_last_page_has_real_data(ws, header_row: int, body_start: int, summary_end: int, real_rows_set: set, page_height_factor: float = 1.0):
    row_height = DEFAULT_TABLE_ROW_HEIGHT
    body_rows = [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden]
    while row_height >= MIN_COMPRESSED_ROW_HEIGHT:
        normalize_table_row_heights(ws, header_row, summary_end, row_height)
        pages, _, _ = paginate_table_rows(ws, body_rows, body_start, page_height_factor)
        if pages and any(row in real_rows_set for row in pages[-1]):
            return row_height
        row_height = round(row_height - 0.25, 2)
    normalize_table_row_heights(ws, header_row, summary_end, MIN_COMPRESSED_ROW_HEIGHT)
    return MIN_COMPRESSED_ROW_HEIGHT


def compress_underfilled_last_page(ws, header_row: int, body_start: int, first_total: int, summary_end: int, real_rows_set: set, page_height_factor: float = 1.0):
    body_rows = [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden]
    pages, used, cap = paginate_table_rows(ws, body_rows, body_start, page_height_factor)
    if len(pages) <= 1:
        return DEFAULT_TABLE_ROW_HEIGHT, False
    last_page = pages[-1]
    last_real_count = sum(1 for row in last_page if row in real_rows_set)
    blank_ratio = max(0.0, (cap - used) / cap) if cap else 0.0
    if last_real_count > UNDERFILLED_LAST_PAGE_REAL_ROWS or blank_ratio < UNDERFILLED_LAST_PAGE_BLANK_RATIO:
        return DEFAULT_TABLE_ROW_HEIGHT, False

    target_pages = len(pages) - 1
    summary_rows = {row for row in range(first_total, summary_end + 1) if not ws.row_dimensions[row].hidden}
    row_height = round(DEFAULT_TABLE_ROW_HEIGHT - 0.25, 2)
    while row_height >= MIN_COMPRESSED_ROW_HEIGHT:
        normalize_table_row_heights(ws, header_row, summary_end, row_height)
        body_rows = [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden]
        new_pages, _, _ = paginate_table_rows(ws, body_rows, body_start, page_height_factor)
        if (
            len(new_pages) <= target_pages
            and new_pages
            and any(row in real_rows_set for row in new_pages[-1])
            and (not summary_rows or summary_rows.issubset(set(new_pages[-1])))
        ):
            return row_height, True
        row_height = round(row_height - 0.25, 2)

    normalize_table_row_heights(ws, header_row, summary_end, DEFAULT_TABLE_ROW_HEIGHT)
    return DEFAULT_TABLE_ROW_HEIGHT, False


def choose_insert_count(ws, body_start: int, first_total: int, summary_end: int, last_real_row: int, real_rows_set: set, max_insert=MAX_AUTO_INSERT_ROWS, page_height_factor: float = 1.0):
    body_rows = [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden]
    pages, used, cap = paginate_table_rows(ws, body_rows, body_start, page_height_factor)
    if not pages or not any(row in real_rows_set for row in pages[-1]):
        return 0
    initial_pages = len(pages)
    insert_pos = sum(1 for row in range(body_start, last_real_row + 1) if not ws.row_dimensions[row].hidden)
    summary_rows = {row for row in range(first_total, summary_end + 1) if not ws.row_dimensions[row].hidden}
    best = 0
    reserved = DEFAULT_TABLE_ROW_HEIGHT * FOOTER_GAP_ROWS
    for count in range(1, max_insert + 1):
        modeled = body_rows[:insert_pos] + [None] * count + body_rows[insert_pos:]
        new_pages, new_used, new_cap = paginate_table_rows(ws, modeled, body_start, page_height_factor)
        if len(new_pages) > initial_pages:
            break
        if not any(row in real_rows_set for row in new_pages[-1]):
            break
        if summary_rows and not summary_rows.issubset(set(new_pages[-1])):
            break
        if new_used <= new_cap - reserved:
            best = count
        else:
            break
    return best


def audit_print_state(ws, header_row: int, first_total: int, summary_end: int, real_rows: list[int], page_height_factor: float = 1.0):
    real_set = set(real_rows)
    body_start = next((row for row in range(header_row + 1, first_total) if not ws.row_dimensions[row].hidden), header_row + 1)
    body_rows = [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden]
    pages, used, cap = paginate_table_rows(ws, body_rows, body_start, page_height_factor)
    issues = []
    if pages:
        last_page = pages[-1]
        if len(pages) > 1 and not any(row in real_set for row in last_page):
            issues.append("last-page-no-detail")
        if len(pages) > 1 and all(is_summary_row(ws, row, effective_table_bounds(ws)) or is_blank_template_row(ws, row, effective_table_bounds(ws)) for row in last_page):
            issues.append("summary-or-blank-only-page")
        summary_rows = {row for row in range(first_total, summary_end + 1) if not ws.row_dimensions[row].hidden}
        if summary_rows and not summary_rows.issubset(set(last_page)):
            issues.append("summary-block-split")
        blank_ratio = max(0.0, (cap - used) / cap) if cap else 0.0
        if blank_ratio > 0.30 and any(row in real_set for row in last_page):
            issues.append(f"last-page-blank-ratio:{blank_ratio:.2f}")
    min_col, _, max_col, _ = print_area_bounds(ws)
    width = 0.0
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        if not ws.column_dimensions[letter].hidden:
            width += ws.column_dimensions[letter].width or 8.43
    if width < 128:
        issues.append(f"visible-width-low:{width:.1f}")
    return {"pages": len(pages), "last_used": round(used, 1), "cap": round(cap, 1), "issues": issues}


def apply_adaptive_row_layout(ws):
    header_row = find_table_header_row(ws)
    if not header_row:
        return {"row_height": None, "inserted": 0, "deleted": 0, "mode": "no-header", "audit": None}
    max_col = effective_table_bounds(ws)
    first_total = find_first_total_row(ws, header_row, max_col)
    if not first_total:
        return {"row_height": None, "inserted": 0, "deleted": 0, "mode": "no-total", "audit": None}
    summary_end = find_summary_end(ws, first_total, max_col)
    body_start = next((row for row in range(header_row + 1, first_total) if not ws.row_dimensions[row].hidden), header_row + 1)
    real_rows = detail_real_rows(ws, header_row, first_total)
    if not real_rows:
        return {"row_height": None, "inserted": 0, "deleted": 0, "mode": "empty", "audit": audit_print_state(ws, header_row, first_total, summary_end, real_rows)}

    normalize_table_row_heights(ws, header_row, summary_end, DEFAULT_TABLE_ROW_HEIGHT)
    mode = "short-insert"
    deleted = 0
    row_height = DEFAULT_TABLE_ROW_HEIGHT
    real_set = set(real_rows)
    page_factor = LONG_TABLE_PAGE_HEIGHT_FACTOR if len(real_rows) >= LONG_TABLE_REAL_ROWS else SHORT_TABLE_PAGE_HEIGHT_FACTOR
    skip_insert = False

    pages, _, _ = paginate_table_rows(ws, [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden], body_start, page_factor)
    if pages and not any(row in real_set for row in pages[-1]):
        mode = "short-trim-insert" if len(real_rows) < LONG_TABLE_REAL_ROWS else "long-trim-compress"
        while True:
            first_total = find_first_total_row(ws, header_row, max_col)
            summary_end = find_summary_end(ws, first_total, max_col)
            blank_rows = [
                row for row in range(real_rows[-1] + 1, first_total)
                if not ws.row_dimensions[row].hidden and is_blank_template_row(ws, row, max_col)
            ]
            if not blank_rows:
                break
            delete_one_row_preserving_formulas(ws, blank_rows[-1])
            deleted += 1
            first_total = find_first_total_row(ws, header_row, max_col)
            summary_end = find_summary_end(ws, first_total, max_col)
            pages, _, _ = paginate_table_rows(
                ws,
                [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden],
                body_start,
                page_factor,
            )
            if pages and any(row in real_set for row in pages[-1]):
                break

    if len(real_rows) >= LONG_TABLE_REAL_ROWS:
        if mode == "short-insert":
            mode = "long-insert"
        while True:
            first_total = find_first_total_row(ws, header_row, max_col)
            summary_end = find_summary_end(ws, first_total, max_col)
            real_rows = detail_real_rows(ws, header_row, first_total)
            blank_rows = [
                row for row in range(real_rows[-1] + 1, first_total)
                if not ws.row_dimensions[row].hidden and is_blank_template_row(ws, row, max_col)
            ]
            if not blank_rows:
                break
            delete_one_row_preserving_formulas(ws, blank_rows[-1])
            deleted += 1
        first_total = find_first_total_row(ws, header_row, max_col)
        summary_end = find_summary_end(ws, first_total, max_col)
        real_rows = detail_real_rows(ws, header_row, first_total)
        real_set = set(real_rows)
        pages, _, _ = paginate_table_rows(ws, [row for row in range(body_start, summary_end + 1) if not ws.row_dimensions[row].hidden], body_start, page_factor)
        if pages and not any(row in set(real_rows) for row in pages[-1]):
            mode = "long-trim-compress"
            while True:
                first_total = find_first_total_row(ws, header_row, max_col)
                summary_end = find_summary_end(ws, first_total, max_col)
                blank_rows = [
                    row for row in range(real_rows[-1] + 1, first_total)
                    if not ws.row_dimensions[row].hidden and is_blank_template_row(ws, row, max_col)
                ]
                if not blank_rows:
                    break
                delete_one_row_preserving_formulas(ws, blank_rows[-1])
                deleted += 1
            first_total = find_first_total_row(ws, header_row, max_col)
            summary_end = find_summary_end(ws, first_total, max_col)
            real_rows = detail_real_rows(ws, header_row, first_total)
            row_height = compress_until_last_page_has_real_data(ws, header_row, body_start, summary_end, set(real_rows), page_factor)
        first_total = find_first_total_row(ws, header_row, max_col)
        summary_end = find_summary_end(ws, first_total, max_col)
        real_rows = detail_real_rows(ws, header_row, first_total)
        compressed_height, compressed = compress_underfilled_last_page(
            ws,
            header_row,
            body_start,
            first_total,
            summary_end,
            set(real_rows),
            page_factor,
        )
        if compressed:
            row_height = compressed_height
            mode = "long-trim-compress"
            skip_insert = True

    first_total = find_first_total_row(ws, header_row, max_col)
    summary_end = find_summary_end(ws, first_total, max_col)
    real_rows = detail_real_rows(ws, header_row, first_total)
    page_factor = LONG_TABLE_PAGE_HEIGHT_FACTOR if len(real_rows) >= LONG_TABLE_REAL_ROWS else SHORT_TABLE_PAGE_HEIGHT_FACTOR
    insert_count = 0 if skip_insert else choose_insert_count(ws, body_start, first_total, summary_end, real_rows[-1], set(real_rows), page_height_factor=page_factor)
    if insert_count:
        insert_at = real_rows[-1] + 1
        template_row = insert_at if insert_at < first_total else real_rows[-1]
        insert_rows_preserving_formulas(ws, insert_at, insert_count, template_row, max_col, row_height)
    first_total = find_first_total_row(ws, header_row, max_col)
    summary_end = find_summary_end(ws, first_total, max_col)
    real_rows = detail_real_rows(ws, header_row, first_total)
    return {
        "row_height": row_height,
        "inserted": insert_count,
        "deleted": deleted,
        "mode": mode,
        "audit": audit_print_state(ws, header_row, first_total, summary_end, real_rows, page_factor),
    }



def clear_evaluator_footer(ws):
    changed = False
    for footer in (ws.oddFooter, ws.evenFooter, ws.firstFooter):
        if footer.right.text is not None:
            footer.right.text = None
            changed = True
    return changed


def tighten_for_signature_area(ws, last_row, limit=620.0):
    before = 0.0
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        before += ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
    if last_row > 40 or before <= limit:
        return before, before
    factor = limit / before
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        old = ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
        floor = 10.8 if old >= 10.8 else old
        ws.row_dimensions[row].height = min(old, max(floor, round(old * factor, 2)))
    after = 0.0
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        after += ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15
    return before, after


def apply_final_print_setup_without_row_tighten(ws, args, base):
    min_col, max_col = base.find_table_bounds(ws)
    last_row = base.find_last_print_row(ws, max_col)
    ws.print_area = f"{get_column_letter(min_col)}1:{get_column_letter(max_col)}{last_row}"
    ws.row_breaks.brk = []
    ws.col_breaks.brk = []
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = args.margin
    ws.page_margins.right = args.margin
    ws.page_margins.top = args.top_margin
    ws.page_margins.bottom = args.bottom_margin
    ws.page_margins.header = args.header_margin
    ws.page_margins.footer = args.footer_margin
    ws.print_options.horizontalCentered = True
    before_h = row_height_sum_until(ws, last_row)
    before_w, after_w = base.resize_visible_cols(ws, args.target_width)
    after_h = row_height_sum_until(ws, last_row)
    return last_row, before_w, after_w, before_h, after_h


def parse_args(argv):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to an appraisal detail .xlsx/.xlsm workbook")
    parser.add_argument("--target-width", type=float, default=143.0)
    parser.add_argument("--margin", type=float, default=0.70)
    parser.add_argument("--top-margin", type=float, default=0.30)
    parser.add_argument("--bottom-margin", type=float, default=0.85)
    parser.add_argument("--header-margin", type=float, default=0.12)
    parser.add_argument("--footer-margin", type=float, default=0.45)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-backup", action="store_true")
    return parser.parse_args(argv)


def make_base_args(args):
    return argparse.Namespace(
        target_width=args.target_width,
        margin=args.margin,
        top_margin=args.top_margin,
        bottom_margin=args.bottom_margin,
        header_margin=args.header_margin,
        footer_margin=args.footer_margin,
        keep_empty_summary=True,
        dry_run=args.dry_run,
        no_backup=True,
    )


def main(argv=None):
    args = parse_args(argv or sys.argv[1:])
    path = Path(args.workbook).expanduser()
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)
    base = load_base_module()
    base_args = make_base_args(args)
    original_sheet_states = {ws.title: ws.sheet_state for ws in wb.worksheets}

    backup = None
    if not args.dry_run and not args.no_backup:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = path.with_name(f"{path.stem}.申报表转换前备份-{stamp}{path.suffix}")
        copy2(path, backup)

    report = []
    for ws in wb.worksheets:
        if "汇总" in ws.title:
            if ws.sheet_state == "visible":
                ws.sheet_state = "hidden"
                report.append((ws.title, "hidden-summary", [], [], False, None))
            continue
        if ws.sheet_state != "visible":
            continue

        cleared_rows, has_real_detail, _, _ = sanitize_and_classify_detail_rows(ws)
        if not has_real_detail:
            ws.sheet_state = "hidden"
            report.append((ws.title, "hidden-empty", [], [], False, None))
            continue

        title_cells = replace_title(ws)
        result_cols = hide_result_columns(ws)
        footer_changed = clear_evaluator_footer(ws)
        clear_fake_rows, deleted_rows, filled_rows = cleared_rows, 0, 0
        summary_blank_rows = 0
        empty_cols = base.hide_empty_columns(ws)
        metrics = base.apply_print_setup(ws, base_args)
        # Fit all visible columns to one printed page width.
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = None
        ws.print_options.horizontalCentered = True
        row_layout = apply_adaptive_row_layout(ws)
        deleted_rows += row_layout["deleted"]
        filled_rows += row_layout["inserted"]
        last_row, before_w, after_w, before_h, after_h = apply_final_print_setup_without_row_tighten(ws, args, base)
        metrics = (last_row, before_w, after_w, before_h, after_h)
        row_changes = []
        if clear_fake_rows:
            row_changes.append(f"clear_fake_rows:{clear_fake_rows}")
        if deleted_rows:
            row_changes.append(f"delete_template_rows:{deleted_rows}")
        if summary_blank_rows:
            row_changes.append(f"delete_summary_blank_rows:{summary_blank_rows}")
        if filled_rows:
            row_changes.append(f"fill_rows:{filled_rows}")
        if 'row_layout' in locals() and row_layout.get("row_height"):
            row_changes.append(f"row_mode:{row_layout['mode']}")
            row_changes.append(f"row_height:{row_layout['row_height']}")
        audit = row_layout.get("audit")
        if audit:
            row_changes.append(f"pages:{audit['pages']}")
            if audit["issues"]:
                row_changes.append(f"audit_issues:{'|'.join(audit['issues'])}")
        report.append((ws.title, "detail", result_cols, title_cells, footer_changed, (empty_cols + row_changes, metrics)))

    if not any(ws.sheet_state == "visible" for ws in wb.worksheets):
        for ws in wb.worksheets:
            ws.sheet_state = original_sheet_states.get(ws.title, ws.sheet_state)
        if not args.dry_run:
            wb.save(path)
        print(f"processed_sheets={len(report)} dry_run={args.dry_run} no_visible_detail=True")
        print("action=no-visible-detail restored_original_sheet_states=True")
        return

    first_visible_index = next(
        index for index, ws in enumerate(wb.worksheets) if ws.sheet_state == "visible"
    )
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = False
    wb.active = first_visible_index
    wb.worksheets[first_visible_index].sheet_view.tabSelected = True

    if not args.dry_run:
        wb.save(path)

    if backup:
        print(f"backup={backup}")
    print(f"processed_sheets={len(report)} dry_run={args.dry_run}")
    for title, action, result_cols, title_cells, footer_changed, detail in report:
        if action in ("hidden-summary", "hidden-empty"):
            print(f"{title}\taction={action}")
            continue
        empty_cols, metrics = detail
        last_row, before_w, after_w, before_h, after_h = metrics
        print(
            f"{title}\taction=detail\ttitle_cells={','.join(title_cells) or '-'}"
            f"\tresult_cols={','.join(result_cols) or '-'}"
            f"\tempty_cols={','.join(empty_cols) or '-'}"
            f"\tfooter_right_cleared={footer_changed}"
            f"\tlast_row={last_row}\twidth={before_w}->{after_w}"
            f"\theight={round(before_h, 1)}->{round(after_h, 1)}"
        )


if __name__ == "__main__":
    main()
